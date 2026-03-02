import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import pandas as pd
import subprocess
from datetime import datetime, timedelta
import openpyxl
import shutil
from tkcalendar import Calendar
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from tkinter.font import Font
import textwrap
from collections import Counter
import glob
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from tkinter import filedialog
from openpyxl import load_workbook
from fpdf import FPDF
from PyPDF2 import PdfReader, PdfWriter
import io
import pyexcel as pe
from difflib import SequenceMatcher
import requests
from collections import defaultdict
from openpyxl.styles import numbers


import builtins

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DADOS_DIR = os.path.join(BASE_DIR, "dados_teste")
DASHBOARD_DIR = os.path.join(DADOS_DIR, "dashboard")
BASE_DADOS_DIR = os.path.join(DADOS_DIR, "base_dados")

LEGACY_DASHBOARD_ROOT = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard"
LEGACY_BASE_DADOS_ROOT = r"G:\ASSISTÊNCIA TÉCNICA\Base de Dados Preventivas e Peças"


def normalizar_caminho(caminho):
    if not isinstance(caminho, str):
        return caminho

    for antigo, novo in (
        (LEGACY_DASHBOARD_ROOT, DASHBOARD_DIR),
        (LEGACY_BASE_DADOS_ROOT, BASE_DADOS_DIR),
    ):
        caminho = caminho.replace(antigo, novo)
        caminho = caminho.replace(antigo.replace("\\", "/"), novo.replace("\\", "/"))

    return os.path.normpath(caminho)


def garantir_estrutura_minima():
    pastas = [
        DADOS_DIR,
        BASE_DADOS_DIR,
        os.path.join(DASHBOARD_DIR, "planilhas"),
        os.path.join(DASHBOARD_DIR, "planilhas", "pecas"),
        os.path.join(DASHBOARD_DIR, "planilhas", "pecas", "cnc"),
        os.path.join(DASHBOARD_DIR, "planilhas", "pecas", "máquina"),
        os.path.join(DASHBOARD_DIR, "planilhas", "pecas", "fonte"),
        os.path.join(DASHBOARD_DIR, "planilhas", "pecas", "acessórios"),
        os.path.join(DASHBOARD_DIR, "planilhas", "historico_de_vendas"),
        os.path.join(DASHBOARD_DIR, "planilhas", "estoque"),
        os.path.join(DASHBOARD_DIR, "planilhas", "clientes_explosao"),
        os.path.join(DASHBOARD_DIR, "planilhas", "informacoes_detalhadas"),
        os.path.join(DASHBOARD_DIR, "planilhas", "atualizacao_preco"),
        os.path.join(DASHBOARD_DIR, "planilhas", "pecas avulsas"),
        os.path.join(DASHBOARD_DIR, "orçamentos", "pendentes"),
        os.path.join(DASHBOARD_DIR, "orçamentos", "cancelados"),
        os.path.join(DASHBOARD_DIR, "orçamentos", "confirmados"),
        os.path.join(DASHBOARD_DIR, "orçamentos", "concluídos"),
        os.path.join(DASHBOARD_DIR, "orçamentos", "historico_orçamentos"),
        os.path.join(DASHBOARD_DIR, "orçamentos", "pdfs"),
        os.path.join(DASHBOARD_DIR, "notificacoes"),
        os.path.join(DASHBOARD_DIR, "personalizacao"),
    ]
    for pasta in pastas:
        os.makedirs(pasta, exist_ok=True)

    arquivos_texto = {
        os.path.join(DASHBOARD_DIR, "planilhas", "motivos_de_cancelamento.txt"): "",
        os.path.join(DASHBOARD_DIR, "personalizacao", "opcionais.txt"): "",
    }
    for caminho, conteudo in arquivos_texto.items():
        if not os.path.exists(caminho):
            with open(caminho, "w", encoding="utf-8") as f:
                f.write(conteudo)

    clientes = os.path.join(DASHBOARD_DIR, "planilhas", "clientes.xlsx")
    if not os.path.exists(clientes):
        pd.DataFrame(columns=["nome_cliente", "segmento", "cidade", "estado"]).to_excel(clientes, index=False)


def _normalizar_argumento_caminho(valor):
    if isinstance(valor, str):
        return normalizar_caminho(valor)
    return valor


_original_open = builtins.open

def _open_compat(arquivo, *args, **kwargs):
    arquivo = _normalizar_argumento_caminho(arquivo)
    if isinstance(arquivo, str) and any(modo in args[0] if args else kwargs.get("mode", "r") for modo in ("w", "a", "x", "+")):
        pasta = os.path.dirname(arquivo)
        if pasta:
            os.makedirs(pasta, exist_ok=True)
    return _original_open(arquivo, *args, **kwargs)

builtins.open = _open_compat

_original_listdir = os.listdir
os.listdir = lambda path: _original_listdir(_normalizar_argumento_caminho(path))

_original_exists = os.path.exists
os.path.exists = lambda path: _original_exists(_normalizar_argumento_caminho(path))

_original_remove = os.remove
os.remove = lambda path: _original_remove(_normalizar_argumento_caminho(path))

_original_makedirs = os.makedirs
os.makedirs = lambda name, mode=0o777, exist_ok=False: _original_makedirs(_normalizar_argumento_caminho(name), mode=mode, exist_ok=exist_ok)

_original_move = shutil.move
shutil.move = lambda src, dst, *a, **k: _original_move(_normalizar_argumento_caminho(src), _normalizar_argumento_caminho(dst), *a, **k)

_original_copy = shutil.copy
shutil.copy = lambda src, dst, *a, **k: _original_copy(_normalizar_argumento_caminho(src), _normalizar_argumento_caminho(dst), *a, **k)

_original_glob = glob.glob
glob.glob = lambda pathname, *a, **k: _original_glob(_normalizar_argumento_caminho(pathname), *a, **k)

_original_read_excel = pd.read_excel
pd.read_excel = lambda io, *a, **k: _original_read_excel(_normalizar_argumento_caminho(io), *a, **k)

_original_read_csv = pd.read_csv
pd.read_csv = lambda filepath_or_buffer, *a, **k: _original_read_csv(_normalizar_argumento_caminho(filepath_or_buffer), *a, **k)

_original_df_to_excel = pd.DataFrame.to_excel

def _to_excel_compat(self, excel_writer, *args, **kwargs):
    excel_writer = _normalizar_argumento_caminho(excel_writer)
    if isinstance(excel_writer, str):
        pasta = os.path.dirname(excel_writer)
        if pasta:
            os.makedirs(pasta, exist_ok=True)
    return _original_df_to_excel(self, excel_writer, *args, **kwargs)

pd.DataFrame.to_excel = _to_excel_compat

if hasattr(os, "startfile"):
    _original_startfile = os.startfile
    os.startfile = lambda path, *a, **k: _original_startfile(_normalizar_argumento_caminho(path), *a, **k)

# função que atualiza tabelas do focco ao iniciar o programa
def converter_csv_para_xlsx():
    # Caminho da pasta onde estão os arquivos CSV
    pasta_origem = r"G:\ASSISTÊNCIA TÉCNICA\Base de Dados Preventivas e Peças"

    # Caminhos de destino
    destino_faturamento = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\planilhas\historico_de_vendas"
    destino_oc = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\planilhas\estoque"
    destino_of = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\planilhas\estoque"
    destino_pv = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\planilhas\estoque"
    destino_saldo = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\planilhas\estoque"

    # Arquivos e destinos
    arquivos_destinos = {
        "FATURAMENTO_PREV_PECAS.csv": destino_faturamento,
        "ITENS_PENDENTES_OC.csv": destino_oc,
        "ITENS_PENDENTES_OF.csv": destino_of,
        "ITENS_PENDENTES_PV.csv": destino_pv,
        "SALDO_ITENS.csv": destino_saldo,
    }

    for arquivo_csv, pasta_destino in arquivos_destinos.items():
        caminho_csv = os.path.join(pasta_origem, arquivo_csv)
        nome_base = os.path.splitext(arquivo_csv)[0]
        caminho_xlsx = os.path.join(pasta_destino, f"{nome_base}.xlsx")

        print("=" * 80)
        print(f"[INFO] Processando arquivo: {arquivo_csv}")
        print(f"[INFO] Caminho CSV esperado: {caminho_csv}")
        print(f"[INFO] Caminho destino XLSX: {caminho_xlsx}")

        if not os.path.exists(caminho_csv):
            print(f"[ERRO] Arquivo CSV não encontrado: {caminho_csv}")
            continue

        os.makedirs(pasta_destino, exist_ok=True)

        # Tenta ler CSV
        try:
            try:
                df = pd.read_csv(caminho_csv, sep=";", encoding="utf-8", dtype=str, engine="python")
                print(f"[OK] CSV lido com sucesso em UTF-8 ({len(df)} linhas).")
            except UnicodeDecodeError:
                print("[WARN] UTF-8 falhou, tentando latin-1...")
                df = pd.read_csv(caminho_csv, sep=";", encoding="latin-1", dtype=str, engine="python")
                print(f"[OK] CSV lido com sucesso em latin-1 ({len(df)} linhas).")

            # Normalizar números
            for col in df.columns:
                try:
                    # Remove espaços e zeros à esquerda, converte vírgula em ponto
                    df[col] = df[col].str.strip().str.replace(',', '.')
                    df[col] = df[col].astype(float)
                except:
                    pass  # se não for numérico, mantém como string

            # Salvar em XLSX
            df.to_excel(caminho_xlsx, index=False, sheet_name="Planilha1")
            print(f"[SUCESSO] Convertido: {arquivo_csv} -> {caminho_xlsx}")

        except Exception as e:
            print(f"[ERRO] Falha ao converter {arquivo_csv}")
            print(f"       Detalhes: {e}")

# Função para criar o estilo do botão
def criar_estilo_botao():
    style = ttk.Style()
    style.configure("TButton",
                    font=('Arial', 12, 'bold'),
                    padding=10,
                    relief="raised",
                    background="#4CAF50",  # Verde
                    foreground="#000000")
    style.map("TButton", background=[('active', '#45a049')])  # Cor de fundo ao passar o mouse
    return style


# Função para criar listas de peças, apagar ou editar alguma lista, podendo abrir o xlsx dela diretamente pelo programa
def abrir_tela_cadastro_pecas():
    tela_cadastro = tk.Toplevel()
    tela_cadastro.title("Cadastro de Peças")
    tela_cadastro.geometry("600x600")
    tela_cadastro.configure(bg="#f4f4f4")
    tela_cadastro.grab_set()


    # Função para exibir ajuda
    def exibir_ajuda_cadastro_pecas():
        mensagem_ajuda = (
            "Explicação das funções na tela de Cadastro de Peças:\n\n"
            "1. **CNC**: Seleciona a pasta de arquivos relacionados ao CNC.\n"
            "2. **Máquina**: Seleciona a pasta de arquivos relacionados à máquina.\n"
            "3. **Fonte**: Seleciona a pasta de arquivos relacionados à fonte.\n"
            "4. **Acessórios**: Seleciona a pasta de arquivos relacionados aos acessórios.\n\n"
            "5. **Abrir Arquivo**: Abre o arquivo selecionado na lista de arquivos.\n"
            "6. **Criar Nova Planilha de Peças**: Cria uma nova planilha para cadastrar novas peças.\n"
            "7. **Excluir Planilha**: Exclui o arquivo selecionado na lista de arquivos.\n"
            "8. **Voltar**: Fecha a tela de cadastro de peças.\n"
        )
        # Cria a tela de ajuda como um "TopLevel", que ficará acima da janela principal
        tela_ajuda = tk.Toplevel()
        tela_ajuda.title("Ajuda - Cadastro de Peças")
        tela_ajuda.geometry("500x300")
        tela_ajuda.configure(bg="#f4f4f4")


        # Exibe a mensagem de ajuda
        message = tk.Label(tela_ajuda, text=mensagem_ajuda, justify="left", padx=10, pady=10)
        message.pack(fill="both", expand=True)

        # Botão para fechar a tela de ajuda
        button_fechar = tk.Button(tela_ajuda, text="Fechar", command=tela_ajuda.destroy)
        button_fechar.pack(pady=10)

    # Ícone de ajuda como botão
    help_button = tk.Button(tela_cadastro, text="?", font=("Arial", 16), command=exibir_ajuda_cadastro_pecas, bg="#f4f4f4", bd=0)
    help_button.grid(row=0, column=4, padx=10, pady=10)  # Posiciona o botão no canto da tela


    # Ícone de ajuda usando um botão com símbolo de interrogação
    help_button = tk.Button(tela_cadastro, text="?", font=("Arial", 16), command=exibir_ajuda_cadastro_pecas, bg="#f4f4f4", bd=0)
    help_button.grid(row=0, column=4, padx=10, pady=10)  # Posiciona o botão no canto da tela

    # Variável global para armazenar a pasta selecionada
    pasta_selecionada = ""

    def listar_arquivos(tipo):
        nonlocal pasta_selecionada
        pasta_selecionada = tipo  # Atualiza a pasta selecionada
        pasta_base = os.path.join("S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\planilhas\\pecas", tipo)  # Diretório fixo
        if os.path.exists(pasta_base):
            arquivos = [f for f in os.listdir(pasta_base) if f.endswith(".xlsx")]
            lista_arquivos.delete(0, tk.END)
            for f in arquivos:
                lista_arquivos.insert(tk.END, f)

    def abrir_arquivo():
        arquivo_selecionado = lista_arquivos.get(tk.ACTIVE)
        if arquivo_selecionado:
            caminho_arquivo = os.path.join("S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\planilhas\\pecas", pasta_selecionada, arquivo_selecionado)
            try:
                os.startfile(caminho_arquivo)
                print(f"Arquivo {arquivo_selecionado} aberto com sucesso.")
            except Exception as e:
                print(f"Erro ao abrir o arquivo: {e}")

    def excluir_planilha():
        arquivo_selecionado = lista_arquivos.get(tk.ACTIVE)
        if arquivo_selecionado:
            resposta = messagebox.askyesno("Excluir Planilha", f"Deseja excluir a planilha {arquivo_selecionado}?")
            if resposta:
                caminho_arquivo = os.path.join("S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\planilhas\\pecas", pasta_selecionada, arquivo_selecionado)
                try:
                    os.remove(caminho_arquivo)
                    lista_arquivos.delete(lista_arquivos.curselection())
                    print(f"Planilha {arquivo_selecionado} excluída com sucesso.")
                except Exception as e:
                    print(f"Erro ao excluir a planilha: {e}")
        else:
            print("Nenhuma planilha selecionada.")

    def criar_nova_planilha_pecas():
        tela_nova_planilha = tk.Toplevel()
        tela_nova_planilha.title("Criar Nova Planilha de Peças")
        tela_nova_planilha.configure(bg="#f4f4f4")
        tela_nova_planilha.grab_set()

        campos = ["Nome da peça", "Quantidade", "Código", "Período de troca", "MO prevista"]
        entradas = []

        ttk.Label(tela_nova_planilha, text="Nome da Planilha", font=('Arial', 12, 'bold')).grid(row=0, column=0,
                                                                                                padx=10, pady=10)
        nome_planilha_entry = ttk.Entry(tela_nova_planilha, font=('Arial', 12))
        nome_planilha_entry.grid(row=0, column=1, padx=10, pady=10)

        frame_entradas = ttk.Frame(tela_nova_planilha)
        frame_entradas.grid(row=1, column=0, columnspan=2)

        def adicionar_novo_item():
            linha = len(entradas)
            nova_entrada = {}
            for i, campo in enumerate(campos):
                ttk.Label(frame_entradas, text=campo, font=('Arial', 10)).grid(row=linha, column=i, padx=5, pady=5)
                entrada = ttk.Entry(frame_entradas, font=('Arial', 10))
                entrada.grid(row=linha, column=i + len(campos), padx=5, pady=5)
                nova_entrada[campo] = entrada
            entradas.append(nova_entrada)

        adicionar_novo_item()

        def salvar_planilha():
            nome_planilha = nome_planilha_entry.get()
            if not nome_planilha:
                print("Nome da planilha não pode ser vazio!")
                return

            dados = []
            for entrada in entradas:
                linha_dados = [entrada["Nome da peça"].get(), entrada["Quantidade"].get(),
                               entrada["Código"].get(), entrada["Período de troca"].get(),
                               entrada["MO prevista"].get()]
                dados.append(linha_dados)

            caminho_arquivo = os.path.join("S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\planilhas\\pecas", pasta_selecionada,
                                           nome_planilha + ".xlsx")

            if not os.path.exists(caminho_arquivo):
                df = pd.DataFrame(dados, columns=campos)
            else:
                df = pd.read_excel(caminho_arquivo)
                for linha in dados:
                    df.loc[len(df)] = linha  # Adiciona cada nova peça à planilha existente

            df.to_excel(caminho_arquivo, index=False)
            tela_nova_planilha.destroy()
            print(f"Planilha {nome_planilha}.xlsx salva com sucesso!")

        ttk.Button(tela_nova_planilha, text="Salvar Planilha", command=salvar_planilha, style="TButton").grid(row=2,
                                                                                                              column=1,
                                                                                                              padx=5,
                                                                                                              pady=10)
        ttk.Button(tela_nova_planilha, text="+", command=adicionar_novo_item, style="TButton").grid(row=2, column=0,
                                                                                                    padx=5, pady=10)

    # Estrutura de layout
    ttk.Button(tela_cadastro, text="CNC", command=lambda: listar_arquivos("CNC"), style="TButton").grid(row=0, column=0,
                                                                                                        padx=10,
                                                                                                        pady=10)
    ttk.Button(tela_cadastro, text="Máquina", command=lambda: listar_arquivos("Maquina"), style="TButton").grid(row=0,
                                                                                                                column=1,
                                                                                                                padx=10,
                                                                                                                pady=10)
    ttk.Button(tela_cadastro, text="Fonte", command=lambda: listar_arquivos("Fonte"), style="TButton").grid(row=0,
                                                                                                            column=2,
                                                                                                            padx=10,
                                                                                                            pady=10)
    ttk.Button(tela_cadastro, text="Acessórios", command=lambda: listar_arquivos("Acessorios"), style="TButton").grid(
        row=0, column=3, padx=10, pady=10)

    lista_arquivos = tk.Listbox(tela_cadastro, font=('Arial', 12))
    lista_arquivos.grid(row=1, column=0, columnspan=4, padx=10, pady=10)

    ttk.Button(tela_cadastro, text="Abrir Arquivo", command=abrir_arquivo, style="TButton").grid(row=2, column=0,
                                                                                                 columnspan=2, padx=10,
                                                                                                 pady=10)
    ttk.Button(tela_cadastro, text="Criar Nova Planilha de Peças", command=criar_nova_planilha_pecas,
               style="TButton").grid(row=2, column=2, columnspan=2, padx=10, pady=10)
    ttk.Button(tela_cadastro, text="Excluir Planilha", command=excluir_planilha, style="TButton").grid(row=3, column=0,
                                                                                                       columnspan=4,
                                                                                                       padx=10, pady=10)
    ttk.Button(tela_cadastro, text="Voltar", command=tela_cadastro.destroy, style="TButton").grid(row=4, column=0,
                                                                                                 columnspan=4, padx=10,
                                                                                                 pady=10)
    root.wait_window(tela_cadastro)

# Função para cadastrar um novo cliente ou uma nova máquina para um cliente, pode apagar a máquina também caso necessário
def abrir_tela_cadastro_clientes():
    tela_clientes = tk.Toplevel()
    tela_clientes.title("Cadastro de Clientes")
    tela_clientes.geometry("700x750")
    tela_clientes.grab_set()

    # Abas
    abas = ttk.Notebook(tela_clientes)
    frame_cadastro = ttk.Frame(abas)
    frame_excluir = ttk.Frame(abas)
    abas.add(frame_cadastro, text="Cadastrar / Editar")
    abas.add(frame_excluir, text="Excluir Cliente/Máquina")
    abas.pack(expand=True, fill="both")


    caminho_arquivo = os.path.join("S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\planilhas\\clientes.xlsx")

    # Funções
    def listar_arquivos(tipo):
        pasta_base = os.path.join("S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\planilhas\\pecas", tipo)
        if os.path.exists(pasta_base):
            arquivos = [f.replace(".xlsx", "") for f in os.listdir(pasta_base) if f.endswith(".xlsx")]
            return sorted(arquivos)
        return []

    def salvar_cliente():
        nome_cliente = entrada_nome.get().strip()
        nome_maquina = entrada_nome_maquina.get().strip()
        modelo_maquina = combo_maquina.get().strip()
        modelo_fonte = combo_fonte.get().strip()
        modelo_cnc = combo_cnc.get().strip()
        modelo_acessorios = combo_acessorios.get().strip()
        distancia = entrada_distancia.get().strip()
        valor_hora = entrada_valor_hora.get().strip()

        if not nome_cliente or not nome_maquina:
            messagebox.showerror("Erro", "Nome do Cliente e Nome da Máquina são obrigatórios.")
            return

        dados_cliente = [nome_cliente, nome_maquina, modelo_maquina, modelo_fonte, modelo_cnc,
                         modelo_acessorios, distancia, valor_hora]

        colunas = ["Nome do Cliente", "Nome da Máquina", "Modelo de Máquina", "Modelo de Fonte",
                   "Modelo de CNC", "Modelo de Acessórios", "Distância (km)", "Valor Mão de Obra (R$/h)"]

        if os.path.exists(caminho_arquivo):
            df = pd.read_excel(caminho_arquivo)

            df_existente = df[(df["Nome do Cliente"] == nome_cliente) & (df["Nome da Máquina"] == nome_maquina)]
            if not df_existente.empty:
                df.loc[df_existente.index, ["Modelo de Máquina", "Modelo de Fonte", "Modelo de CNC",
                                            "Modelo de Acessórios", "Distância (km)", "Valor Mão de Obra (R$/h)"]] = [
                    modelo_maquina, modelo_fonte, modelo_cnc, modelo_acessorios,
                    distancia, valor_hora
                ]
                mensagem = "Cliente/Máquina atualizados com sucesso!"
            else:
                df.loc[len(df)] = dados_cliente
                mensagem = "Novo Cliente/Máquina cadastrados com sucesso!"

            df.to_excel(caminho_arquivo, index=False)
        else:
            df = pd.DataFrame([dados_cliente], columns=colunas)
            df.to_excel(caminho_arquivo, index=False)
            mensagem = "Primeiro Cliente/Máquina cadastrados com sucesso!"

        messagebox.showinfo("Sucesso", mensagem)
        limpar_campos()

    def limpar_campos():
        entrada_nome.delete(0, tk.END)
        entrada_nome_maquina.delete(0, tk.END)
        combo_maquina.set('')
        combo_fonte.set('')
        combo_cnc.set('')
        combo_acessorios.set('')
        entrada_distancia.delete(0, tk.END)
        entrada_valor_hora.delete(0, tk.END)
        entrada_nome.focus()

    def filtrar_clientes(event):
        texto = entrada_nome.get().lower()
        listbox_clientes.delete(0, tk.END)

        if os.path.exists(caminho_arquivo):
            df = pd.read_excel(caminho_arquivo)
            clientes_filtrados = df[df["Nome do Cliente"].str.contains(texto, case=False, na=False)]["Nome do Cliente"].unique()
            for cliente in sorted(clientes_filtrados):
                listbox_clientes.insert(tk.END, cliente)

    def selecionar_cliente(event):
        if listbox_clientes.curselection():
            nome_cliente = listbox_clientes.get(listbox_clientes.curselection())
            entrada_nome.delete(0, tk.END)
            entrada_nome.insert(0, nome_cliente)
            listar_maquinas_cliente(nome_cliente)

    def listar_maquinas_cliente(nome_cliente):
        if os.path.exists(caminho_arquivo):
            df = pd.read_excel(caminho_arquivo)
            maquinas = df[df["Nome do Cliente"] == nome_cliente]["Nome da Máquina"].dropna().unique()

            if len(maquinas) > 1:
                janela_maquinas = tk.Toplevel()
                janela_maquinas.title("Selecionar Máquina")
                janela_maquinas.geometry("300x150")
                janela_maquinas.grab_set()

                ttk.Label(janela_maquinas, text="Selecione o nome da máquina:").pack(pady=10)
                combo_maquinas_cliente = ttk.Combobox(janela_maquinas, values=sorted(maquinas))
                combo_maquinas_cliente.pack(pady=5)

                def confirmar_maquina():
                    nome_maquina = combo_maquinas_cliente.get()
                    entrada_nome_maquina.delete(0, tk.END)
                    entrada_nome_maquina.insert(0, nome_maquina)
                    carregar_informacoes_cliente(nome_cliente, nome_maquina)
                    janela_maquinas.destroy()

                ttk.Button(janela_maquinas, text="Confirmar", command=confirmar_maquina).pack(pady=10)
            else:
                nome_maquina = maquinas[0] if len(maquinas) == 1 else ""
                entrada_nome_maquina.delete(0, tk.END)
                entrada_nome_maquina.insert(0, nome_maquina)
                carregar_informacoes_cliente(nome_cliente, nome_maquina)

    def carregar_informacoes_cliente(nome_cliente, nome_maquina):
        if os.path.exists(caminho_arquivo):
            df = pd.read_excel(caminho_arquivo)
            cliente_encontrado = df[(df["Nome do Cliente"] == nome_cliente) & (df["Nome da Máquina"] == nome_maquina)]

            if not cliente_encontrado.empty:
                combo_maquina.set(cliente_encontrado.iloc[0]["Modelo de Máquina"])
                combo_fonte.set(cliente_encontrado.iloc[0]["Modelo de Fonte"])
                combo_cnc.set(cliente_encontrado.iloc[0]["Modelo de CNC"])
                combo_acessorios.set(cliente_encontrado.iloc[0]["Modelo de Acessórios"])
                entrada_distancia.delete(0, tk.END)
                entrada_distancia.insert(0, str(cliente_encontrado.iloc[0].get("Distância (km)", "")))
                entrada_valor_hora.delete(0, tk.END)
                entrada_valor_hora.insert(0, str(cliente_encontrado.iloc[0].get("Valor Mão de Obra (R$/h)", "")))

    def excluir_cliente():
        nome_cliente = entrada_excluir_cliente.get().strip()
        nome_maquina = entrada_excluir_maquina.get().strip()

        if not nome_cliente or not nome_maquina:
            messagebox.showerror("Erro", "Preencha o Nome do Cliente e Nome da Máquina para excluir.")
            return

        if os.path.exists(caminho_arquivo):
            df = pd.read_excel(caminho_arquivo)
            df_filtrado = df[~((df["Nome do Cliente"] == nome_cliente) & (df["Nome da Máquina"] == nome_maquina))]

            if len(df) == len(df_filtrado):
                messagebox.showinfo("Aviso", "Cliente/Máquina não encontrados.")
            else:
                df_filtrado.to_excel(caminho_arquivo, index=False)
                messagebox.showinfo("Sucesso", "Cliente/Máquina excluídos com sucesso!")
                entrada_excluir_cliente.delete(0, tk.END)
                entrada_excluir_maquina.delete(0, tk.END)

    def filtrar_clientes_excluir(event):
        texto = entrada_excluir_cliente.get().lower()
        listbox_clientes_excluir.delete(0, tk.END)

        if os.path.exists(caminho_arquivo):
            df = pd.read_excel(caminho_arquivo)
            clientes_filtrados = df[df["Nome do Cliente"].str.contains(texto, case=False, na=False)]["Nome do Cliente"].unique()
            for cliente in sorted(clientes_filtrados):
                listbox_clientes_excluir.insert(tk.END, cliente)

    def selecionar_cliente_excluir(event):
        if listbox_clientes_excluir.curselection():
            nome_cliente = listbox_clientes_excluir.get(listbox_clientes_excluir.curselection())
            entrada_excluir_cliente.delete(0, tk.END)
            entrada_excluir_cliente.insert(0, nome_cliente)
            listar_maquinas_excluir(nome_cliente)

    def listar_maquinas_excluir(nome_cliente):
        if os.path.exists(caminho_arquivo):
            df = pd.read_excel(caminho_arquivo)
            maquinas = df[df["Nome do Cliente"] == nome_cliente]["Nome da Máquina"].dropna().unique()
            listbox_maquinas_excluir.delete(0, tk.END)
            for maquina in sorted(maquinas):
                listbox_maquinas_excluir.insert(tk.END, maquina)

    def selecionar_maquina_excluir(event):
        if listbox_maquinas_excluir.curselection():
            nome_maquina = listbox_maquinas_excluir.get(listbox_maquinas_excluir.curselection())
            entrada_excluir_maquina.delete(0, tk.END)
            entrada_excluir_maquina.insert(0, nome_maquina)

    # Frame Cadastro
    ttk.Label(frame_cadastro, text="Nome do Cliente:").pack(pady=5)
    entrada_nome = ttk.Entry(frame_cadastro)
    entrada_nome.pack(pady=5)
    entrada_nome.bind("<KeyRelease>", filtrar_clientes)

    ttk.Label(frame_cadastro, text="Nome da Máquina:").pack(pady=5)
    entrada_nome_maquina = ttk.Entry(frame_cadastro)
    entrada_nome_maquina.pack(pady=5)

    listbox_clientes = tk.Listbox(frame_cadastro, height=5)
    listbox_clientes.pack(pady=5)
    listbox_clientes.bind("<<ListboxSelect>>", selecionar_cliente)

    ttk.Label(frame_cadastro, text="Modelo de Máquina:").pack(pady=5)
    combo_maquina = ttk.Combobox(frame_cadastro, values=listar_arquivos("maquina"))
    combo_maquina.pack(pady=5)

    ttk.Label(frame_cadastro, text="Modelo de Fonte:").pack(pady=5)
    combo_fonte = ttk.Combobox(frame_cadastro, values=listar_arquivos("fonte"))
    combo_fonte.pack(pady=5)

    ttk.Label(frame_cadastro, text="Modelo de CNC:").pack(pady=5)
    combo_cnc = ttk.Combobox(frame_cadastro, values=listar_arquivos("cnc"))
    combo_cnc.pack(pady=5)

    ttk.Label(frame_cadastro, text="Modelo de Acessórios:").pack(pady=5)
    combo_acessorios = ttk.Combobox(frame_cadastro, values=listar_arquivos("acessorios"))
    combo_acessorios.pack(pady=5)

    ttk.Label(frame_cadastro, text="Distância (km) - Opcional:").pack(pady=5)
    entrada_distancia = ttk.Entry(frame_cadastro)
    entrada_distancia.pack(pady=5)

    ttk.Label(frame_cadastro, text="Valor da Mão de Obra (R$/h) - Opcional:").pack(pady=5)
    entrada_valor_hora = ttk.Entry(frame_cadastro)
    entrada_valor_hora.pack(pady=5)

    frame_botoes_cadastro = ttk.Frame(frame_cadastro)
    frame_botoes_cadastro.pack(pady=20)

    ttk.Button(frame_botoes_cadastro, text="Salvar Cliente", command=salvar_cliente).pack(side=tk.LEFT, padx=10)
    ttk.Button(frame_botoes_cadastro, text="Voltar", command=tela_clientes.destroy).pack(side=tk.RIGHT, padx=10)

    # Frame Excluir
    ttk.Label(frame_excluir, text="Nome do Cliente:").pack(pady=5)
    entrada_excluir_cliente = ttk.Entry(frame_excluir)
    entrada_excluir_cliente.pack(pady=5)
    entrada_excluir_cliente.bind("<KeyRelease>", filtrar_clientes_excluir)

    ttk.Label(frame_excluir, text="Nome da Máquina:").pack(pady=5)
    entrada_excluir_maquina = ttk.Entry(frame_excluir)
    entrada_excluir_maquina.pack(pady=5)

    listbox_clientes_excluir = tk.Listbox(frame_excluir, height=5)
    listbox_clientes_excluir.pack(pady=5)
    listbox_clientes_excluir.bind("<<ListboxSelect>>", selecionar_cliente_excluir)

    listbox_maquinas_excluir = tk.Listbox(frame_excluir, height=5)
    listbox_maquinas_excluir.pack(pady=5)
    listbox_maquinas_excluir.bind("<<ListboxSelect>>", selecionar_maquina_excluir)

    ttk.Button(frame_excluir, text="Excluir Cliente/Máquina", command=excluir_cliente).pack(pady=20)



    # Foco inicial
    entrada_nome.focus()


# Função para gerar orçamento, selecionar as peças para o orçamento
def abrir_tela_orcamento_preventivas():
    tela_orcamento = tk.Toplevel()
    tela_orcamento.title("Orçamento de Preventivas")
    tela_orcamento.geometry("600x600")
    tela_orcamento.grab_set()

    def carregar_clientes():
        caminho_clientes = os.path.join(
            "S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\planilhas\\clientes.xlsx")
        if os.path.exists(caminho_clientes):
            df = pd.read_excel(caminho_clientes, header=None)
            return df[0][1:].tolist(), df
        return [], pd.DataFrame()

    def carregar_maquinas(cliente, df_clientes):
        cliente_info = df_clientes[df_clientes[0] == cliente]
        if cliente_info.empty:
            return []
        maquinas = cliente_info[1].dropna().unique().tolist()
        print(f"Máquinas carregadas para o cliente {cliente}: {maquinas}")
        return maquinas

    def filtrar_clientes(event):
        texto_digitado = entrada_cliente.get().lower()
        lista_filtrada = [c for c in clientes if texto_digitado in c.lower()]
        combo_cliente["values"] = lista_filtrada

    def selecionar_cliente(event):
        cliente_selecionado = combo_cliente.get()
        if cliente_selecionado:
            # Carregar as máquinas para o cliente selecionado
            maquinas = carregar_maquinas(cliente_selecionado, df_clientes)
            combo_maquina["values"] = maquinas
            if maquinas:
                maquina_selecionada = combo_maquina.get()  # Pega a máquina selecionada atualmente
                print(f"Máquina selecionada para o cliente {cliente_selecionado}: {maquina_selecionada}")


    def selecionar_maquina(event):
        cliente_selecionado = combo_cliente.get()
        maquina_selecionada = combo_maquina.get()  # Pega a máquina selecionada no combo
        print(f"Máquina selecionada no evento: {maquina_selecionada}")  # Verificando qual máquina foi selecionada
        if cliente_selecionado and maquina_selecionada:
            carregar_pecas(cliente_selecionado, maquina_selecionada)

    # Função que abre tela com todas as peças para serem selecionadas
    class ToolTip:
        """Cria tooltip para qualquer widget"""

        def __init__(self, widget, text):
            self.widget = widget
            self.text = text
            self.tip_window = None
            widget.bind("<Enter>", self.show)
            widget.bind("<Leave>", self.hide)

        def show(self, event=None):
            if self.tip_window or not self.text:
                return
            x, y, _, cy = self.widget.bbox("insert")
            x = x + self.widget.winfo_rootx() + 25
            y = y + cy + self.widget.winfo_rooty() + 20
            self.tip_window = tw = tk.Toplevel(self.widget)
            tw.wm_overrideredirect(True)
            tw.wm_geometry(f"+{x}+{y}")
            label = tk.Label(tw, text=self.text, background="#ffffe0", relief="solid", borderwidth=1,
                             font=("Arial", 9))
            label.pack(ipadx=5, ipady=2)

        def hide(self, event=None):
            if self.tip_window:
                self.tip_window.destroy()
            self.tip_window = None

    def abrir_tela_selecao_pecas(cliente, pecas):
        tela_pecas = tk.Toplevel()
        tela_pecas.title("Seleção de Peças")
        tela_pecas.geometry("1000x600")

        tela_orcamento.destroy()

        pecas_checkboxes = []

        # Frame superior para filtros
        filtro_frame = ttk.Frame(tela_pecas)
        filtro_frame.pack(fill=tk.X, pady=5, padx=5)

        ttk.Label(filtro_frame, text="Filtrar por Período de Troca:").pack(side=tk.LEFT, padx=5)
        periodos = sorted(list({str(p[3]) for p in pecas}))
        periodos.insert(0, "Todos")
        filtro_var = tk.StringVar(value="Todos")
        filtro_combo = ttk.Combobox(filtro_frame, values=periodos, textvariable=filtro_var, state="readonly")
        filtro_combo.pack(side=tk.LEFT, padx=5)

        filtro_estoque_var = tk.IntVar()
        ttk.Checkbutton(filtro_frame, text="Apenas com estoque disponível", variable=filtro_estoque_var).pack(
            side=tk.LEFT, padx=10)

        # Frame principal com canvas e scrollbar
        frame_lista = ttk.Frame(tela_pecas)
        frame_lista.pack(fill=tk.BOTH, expand=True)

        canvas = tk.Canvas(frame_lista, bg="white")
        scrollbar = ttk.Scrollbar(frame_lista, orient=tk.VERTICAL, command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        scrollable_frame = ttk.Frame(canvas)
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        scrollable_frame.bind_all("<MouseWheel>", _on_mousewheel)

        # Carregar planilhas de estoque
        caminho_estoque = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\planilhas\estoque\SALDO_ITENS.xlsx"
        caminho_pv = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\planilhas\estoque\ITENS_PENDENTES_PV.xlsx"
        caminho_of = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\planilhas\estoque\ITENS_PENDENTES_OF.xlsx"

        df_saldo = pd.read_excel(caminho_estoque, header=None)
        df_pv = pd.read_excel(caminho_pv, header=None)
        df_of = pd.read_excel(caminho_of, header=None)

        # Carregar histórico de faturamento
        caminho_faturamento = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\planilhas\historico_de_vendas\FATURAMENTO_PREV_PECAS.xlsx"
        df_faturamento = pd.read_excel(caminho_faturamento, header=None)

        # Campo para exibir o total de peças e serviço
        total_pecas_var = tk.DoubleVar(value=0.0)
        total_servico_var = tk.DoubleVar(value=0.0)

        ttk.Label(tela_pecas, text="Total Peças:").pack(side=tk.LEFT, padx=5)
        ttk.Label(tela_pecas, textvariable=total_pecas_var).pack(side=tk.LEFT, padx=5)

        ttk.Label(tela_pecas, text=" | Serviço Estimado:").pack(side=tk.LEFT, padx=5)
        ttk.Label(tela_pecas, textvariable=total_servico_var).pack(side=tk.LEFT, padx=5)

        precos_manuais = {}

        # Carregar planilha de clientes
        caminho_clientes = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\planilhas\clientes.xlsx"
        df_clientes = pd.read_excel(caminho_clientes, header=None)

        # Buscar dados do cliente (distância e mão de obra)
        cliente_row = df_clientes[df_clientes.iloc[:, 0].astype(str).str.strip() == str(cliente).strip()]
        if not cliente_row.empty:
            distancia_cliente = float(cliente_row.iloc[0, 6])
            mao_obra_cliente = float(cliente_row.iloc[0, 7])
        else:
            distancia_cliente = 0
            mao_obra_cliente = 0



        def calcular_estoque_disponiveis(codigo):
            codigo = str(codigo).strip()
            registros = df_saldo[df_saldo.iloc[:, 1].astype(str).str.strip() == codigo]
            estoque = int(registros.iloc[0, 3]) if not registros.empty else 0
            pendentes_pv = df_pv[df_pv.iloc[:, 0].astype(str).str.strip() == codigo]
            pendentes_total = pendentes_pv.iloc[:, 2].sum() if not pendentes_pv.empty else 0
            pendentes_of = df_of[df_of.iloc[:, 0].astype(str).str.strip() == codigo]
            pendentes_total += pendentes_of.iloc[:, 2].sum() if not pendentes_of.empty else 0
            disponiveis = max(estoque - pendentes_total, 0)
            return estoque, disponiveis

        def atualizar_total():
            total_pecas = 0
            total_servico = 0

            # ---- PEÇAS ----
            for var, peca in pecas_checkboxes:
                if var.get() == 1:
                    codigo = str(peca[2]).strip()
                    qtd = int(peca[1]) if str(peca[1]).isdigit() else 1

                    registros = df_faturamento[df_faturamento.iloc[:, 5].astype(str).str.strip() == codigo]
                    if not registros.empty:
                        preco = registros.iloc[-1, 9]
                        total_pecas += preco * qtd
                    else:
                        if codigo in precos_manuais:
                            total_pecas += precos_manuais[codigo] * qtd
                        else:
                            preco_manual = simpledialog.askfloat(
                                "Valor não encontrado",
                                f"Valor de venda para o código {codigo} não encontrado. Insira manualmente:"
                            )
                            if preco_manual:
                                precos_manuais[codigo] = preco_manual
                                total_pecas += preco_manual * qtd

                    # ---- SERVIÇO POR PEÇA (MO × mão de obra cliente) ----
                    try:
                        mo_item = float(peca[4])  # índice 4 = coluna MO
                    except:
                        mo_item = 0
                    total_servico += mo_item * mao_obra_cliente

            # ---- DESLOCAMENTO ----
            deslocamento = distancia_cliente * 1.5
            tempo_horas = distancia_cliente / 80
            custo_tempo = tempo_horas * (mao_obra_cliente / 2)

            total_servico += deslocamento + custo_tempo

            # Atualizar variáveis de tela
            total_pecas_var.set(total_pecas)
            total_servico_var.set(total_servico)

        # Renderizar peças
        def renderizar_pecas(*args):
            for widget in scrollable_frame.winfo_children():
                widget.destroy()

            headers = ["Selecionar", "Descrição", "Qtd", "Código", "Período", "MO", "Estoque", "Disponíveis"]
            for col, header in enumerate(headers):
                lbl = ttk.Label(scrollable_frame, text=header, font=("Arial", 12, "bold"), background="#cfe2f3")
                lbl.grid(row=0, column=col, sticky="nsew", padx=3, pady=3)
                scrollable_frame.grid_columnconfigure(col, weight=1)

            pecas_checkboxes.clear()
            row_index = 1

            # pega valores dos filtros
            filtro_periodo = filtro_var.get()
            apenas_com_estoque = filtro_estoque_var.get()

            for peca in pecas:
                estoque, disponiveis = calcular_estoque_disponiveis(str(peca[2]))

                # ---- APLICA OS FILTROS ----
                if filtro_periodo != "Todos" and str(peca[3]) != filtro_periodo:
                    continue
                if apenas_com_estoque and disponiveis == 0:
                    continue
                # ---------------------------

                var = tk.IntVar()
                pecas_checkboxes.append((var, peca))
                row_color = '#f9f9f9' if row_index % 2 == 0 else 'white'

                cb = ttk.Checkbutton(scrollable_frame, variable=var, command=atualizar_total)
                cb.grid(row=row_index, column=0, sticky="w", padx=5, pady=3)

                ttk.Label(scrollable_frame, text=peca[0], background=row_color).grid(row=row_index, column=1,
                                                                                     sticky="w", padx=5)
                ttk.Label(scrollable_frame, text=str(peca[1]), background=row_color).grid(row=row_index, column=2,
                                                                                          sticky="w", padx=5)
                ttk.Label(scrollable_frame, text=str(peca[2]), background=row_color).grid(row=row_index, column=3,
                                                                                          sticky="w", padx=5)
                ttk.Label(scrollable_frame, text=str(peca[3]), background=row_color).grid(row=row_index, column=4,
                                                                                          sticky="w", padx=5)
                ttk.Label(scrollable_frame, text=str(peca[4]), background=row_color).grid(row=row_index, column=5,
                                                                                          sticky="w", padx=5)

                cor_estoque = "#ffcccc" if disponiveis == 0 else row_color
                ttk.Label(scrollable_frame, text=str(estoque), background=cor_estoque).grid(row=row_index, column=6,
                                                                                            sticky="w", padx=5)
                ttk.Label(scrollable_frame, text=str(disponiveis), background=cor_estoque).grid(row=row_index, column=7,
                                                                                                sticky="w", padx=5)

                row_index += 1

        # Vincula os filtros à função
        filtro_combo.bind("<<ComboboxSelected>>", renderizar_pecas)
        filtro_estoque_var.trace_add("write", lambda *a: renderizar_pecas())

        # Render inicial
        renderizar_pecas()

        filtro_var.trace("w", renderizar_pecas)
        filtro_estoque_var.trace("w", renderizar_pecas)
        renderizar_pecas()

        # Atualizar tabela automaticamente ao mudar o filtro
        filtro_var.trace("w", renderizar_pecas)
        filtro_estoque_var.trace("w", renderizar_pecas)

        renderizar_pecas()

        # Botões
        botoes_frame = ttk.Frame(tela_pecas)
        botoes_frame.pack(fill=tk.X, pady=5)
        ttk.Button(botoes_frame, text="Gerar Orçamento", command=lambda: gerar_orcamento(pecas_checkboxes)).pack(
            side=tk.LEFT, padx=10)

        def gerar_orcamento(pecas_checkboxes):
            # Pegando apenas as peças marcadas
            pecas_marcadas = [p[1] for p in pecas_checkboxes if p[0].get() == 1]

            if not pecas_marcadas:
                messagebox.showwarning("Aviso", "Nenhuma peça selecionada!")
                return

            # --- NOVO: verificar se as peças existem no arquivo do cliente ---
            if not verificar_pecas_no_arquivo(cliente, pecas_marcadas):
                # Se a verificação retornar False, interrompe a geração do orçamento
                return

            print("[INÍCIO] Baixando planilha de atualização de preços...")
            url = "https://brbaw-my.sharepoint.com/:x:/g/personal/fernando_silva_grupobaw_com_br/ESJ9BKFvMGxcnOxiU87gFLgBZb2sWtl-1cShhaQubjogVw?download=1"
            pasta_atualizacao = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\planilhas\atualizacao_preco"
            os.makedirs(pasta_atualizacao, exist_ok=True)
            caminho_planilha_atualizada = os.path.join(pasta_atualizacao, "consulta_precos_atualizada.xlsx")

            try:
                response = requests.get(url)
                response.raise_for_status()
                with open(caminho_planilha_atualizada, 'wb') as f:
                    f.write(response.content)
                print(f"[SUCESSO] Planilha de atualização salva em: {caminho_planilha_atualizada}")
            except Exception as e:
                print(f"[ERRO] Falha ao baixar ou salvar a planilha de atualização: {e}")
                return

            caminho_orcamentos = os.path.join(
                r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\orçamentos\pendentes"
            )
            os.makedirs(caminho_orcamentos, exist_ok=True)
            arquivos_existentes = os.listdir(caminho_orcamentos)
            numeros_existentes = [
                int(f.split(" ")[1])
                for f in arquivos_existentes
                if f.startswith("Orçamento ") and f.split(" ")[1].isdigit()
            ]
            proximo_id = max(numeros_existentes, default=0) + 1

            partes = cliente.split(" - ")
            nome_cliente = partes[0] if len(partes) > 0 else "Desconhecido"
            nome_maquina = partes[1] if len(partes) > 1 else "Desconhecida"

            nome_base = f"Orçamento {proximo_id} - {nome_cliente} - {pd.Timestamp.now().strftime('%Y-%m-%d')}.xlsx"
            caminho_arquivo_padrao = os.path.join(caminho_orcamentos, nome_base)

            # Criar DataFrame com peças marcadas
            df_orcamento = pd.DataFrame(
                pecas_marcadas,
                columns=["Nome", "Quantidade", "Código", "Período de Troca", "MO Prevista"]
            )
            df_orcamento["Cliente"] = nome_cliente
            df_orcamento["Máquina"] = nome_maquina

            # Calcular estoque e disponíveis para cada peça
            saldos_totais = []
            saldos_disponiveis = []
            for codigo in df_orcamento["Código"]:
                estoque, disponiveis = calcular_estoque_disponiveis(codigo)
                saldos_totais.append(estoque)
                saldos_disponiveis.append(disponiveis)

            df_orcamento["Saldo total"] = saldos_totais
            df_orcamento["Saldo disponível"] = saldos_disponiveis

            # Verificar atualização de preços
            try:
                df_atualizacao = pd.read_excel(caminho_planilha_atualizada, sheet_name=0)
                status_atualizacao = []
                for codigo in df_orcamento["Código"]:
                    linha = df_atualizacao[df_atualizacao.iloc[:, 0] == codigo]
                    if not linha.empty:
                        status = linha.iloc[0, 3]  # Coluna D
                        if isinstance(status, str) and status.strip().upper() == "ATUALIZADO":
                            status_atualizacao.append("ATUALIZADO")
                        else:
                            status_atualizacao.append("SOLICITAR ATUALIZAÇÃO")
                    else:
                        status_atualizacao.append("CÓDIGO NÃO ENCONTRADO")
                df_orcamento["Status Atualização"] = status_atualizacao
                print(f"[OK] Status de atualização verificado para {len(status_atualizacao)} itens.")
            except Exception as e:
                print(f"[ERRO] Erro ao verificar atualização de preços: {e}")
                df_orcamento["Status Atualização"] = "ERRO NA CONSULTA"

            # Criar colunas vazias caso não existam
            colunas_novas = ["Data Início", "Data Fim", "Valor das Peças", "Valor Mão de Obra"]
            for col in colunas_novas:
                if col not in df_orcamento.columns:
                    df_orcamento[col] = ""

            # Reorganizar apenas as colunas a partir de "Máquina"
            col_pos = df_orcamento.columns.get_loc("Máquina")
            col_reordenar = [
                "Máquina",
                "Status Atualização",
                "Data Início",
                "Data Fim",
                "Valor das Peças",
                "Valor Mão de Obra",
                "Saldo total",
                "Saldo disponível"
            ]
            # Mantém todas as colunas anteriores e reorganiza o restante
            col_final = list(df_orcamento.columns[:col_pos]) + col_reordenar
            df_orcamento = df_orcamento[col_final]

            # Salvar orçamento
            df_orcamento.loc["Total", "MO Prevista"] = df_orcamento["MO Prevista"].sum()
            df_orcamento.to_excel(caminho_arquivo_padrao, index=False)

            caminho_personalizado = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Planilhas Excel", "*.xlsx")],
                initialfile=nome_base,
                title="Salvar orçamento como..."
            )

            if caminho_personalizado:
                df_orcamento.to_excel(caminho_personalizado, index=False)
                os.startfile(caminho_personalizado)
            else:
                os.startfile(caminho_arquivo_padrao)

            tela_pecas.destroy()



    def verificar_pecas_no_arquivo(cliente, pecas_marcadas):
        caminho_pasta = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\planilhas\clientes_explosao"
        nome_arquivo_cliente = f"{cliente}.xls"
        caminho_arquivo = os.path.join(caminho_pasta, nome_arquivo_cliente)

        if not os.path.exists(caminho_arquivo):
            messagebox.showinfo(
                "Arquivo não encontrado",
                f"Arquivo do cliente '{cliente}' não foi encontrado. Selecione manualmente."
            )
            caminho_arquivo = filedialog.askopenfilename(
                title="Selecionar arquivo do cliente",
                filetypes=[("Arquivos Excel", "*.xls;*.xlsx")]
            )
            if not caminho_arquivo:
                messagebox.showerror("Erro", "Nenhum arquivo selecionado.")
                return None
            shutil.copy(caminho_arquivo, os.path.join(caminho_pasta, nome_arquivo_cliente))
            caminho_arquivo = os.path.join(caminho_pasta, nome_arquivo_cliente)

        try:
            if caminho_arquivo.endswith(".xls"):
                tabelas = pd.read_html(caminho_arquivo)
                df = tabelas[0]
            else:
                df = pd.read_excel(caminho_arquivo)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao abrir o arquivo: {e}")
            return None

        # Garantir que a coluna E (índice 4) seja string
        df.iloc[:, 4] = df.iloc[:, 4].astype(str)

        resultado = []

        for peca in pecas_marcadas:
            descricao = peca[1]
            codigo = str(peca[2])
            registros = df[df.iloc[:, 4] == codigo]  # Coluna E

            if not registros.empty:
                # Converter para numérico corretamente
                col_qtd = registros.iloc[:, 7].astype(str) \
                    .str.replace(',', '.') \
                    .str.extract(r'([\d\.]+)')[0]
                col_qtd = pd.to_numeric(col_qtd, errors='coerce').fillna(0)
                qtd_total = col_qtd.sum()
                resultado.append(f"CÓDIGO {codigo} | DESC: {descricao} → ENCONTRADO (QTD: {qtd_total})")
            else:
                resultado.append(f"CÓDIGO {codigo} | DESC: {descricao} → NÃO ENCONTRADO")

        resposta = messagebox.askyesno(
            "Resultado da Verificação",
            "\n".join(resultado) + "\n\nDeseja continuar com o orçamento?"
        )

        if not resposta:
            return False

        return True

    # Função para calcular similaridade entre nomes
    # Função para calcular similaridade entre nomes
    def similaridade(a, b):
        return SequenceMatcher(None, a.lower(), b.lower()).ratio()


    # Função para normalizar quantidade
    def normalizar_quantidade(valor):
        if pd.isna(valor):
            return 0
        # Converte para string
        s = str(valor).strip()
        # Remove todos os caracteres que não sejam dígitos
        s = ''.join(c for c in s if c.isdigit())
        # Remove todos os zeros
        s = s.replace('0', '')
        # Se ficar vazio, retorna 0
        return int(s) if s else 0

    # Função principal para carregar peças
    def carregar_pecas(cliente, maquina):
        print(f"\n[INÍCIO] Carregando peças para: Cliente: {cliente}, Máquina: {maquina}")

        caminho_excel = os.path.join(
            r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\planilhas\historico_de_vendas",
            "FATURAMENTO_PREV_PECAS.xlsx"
        )

        pecas_compradas_ano = []
        print("[DEBUG] Tentando abrir:", caminho_excel)

        if not os.path.exists(caminho_excel):
            print("[ERRO] Arquivo não encontrado pelo Python.")
            print("[DEBUG] Lista de arquivos nessa pasta:")
            print(os.listdir(os.path.dirname(caminho_excel)))
            return

        print("[INFO] Arquivo .xlsx encontrado. Lendo com pandas...")

        try:
            df_historico = pd.read_excel(caminho_excel, header=None)
            print(f"[INFO] Total de linhas lidas: {len(df_historico)}")

            data_atual = datetime.now()
            ano_passado = data_atual - timedelta(days=365)

            print("[INFO] Iniciando busca por peças compradas no último ano...")

            for index, row in df_historico.iterrows():
                try:
                    nome_cliente = str(row[3]).strip()  # Coluna D (índice 3)
                    score = similaridade(cliente, nome_cliente)
                    if score >= 0.9:
                        data_venda = pd.to_datetime(row[1], errors='coerce')  # Coluna B (índice 1)
                        if pd.notna(data_venda) and data_venda >= ano_passado:
                            codigo = str(row[5]).strip()  # Coluna f (índice 5)
                            quantidade = normalizar_quantidade(row[7])  # Coluna H (índice 7)
                            descricao = str(row[6]).strip()  # Coluna G (índice 6)

                            pecas_compradas_ano.append(
                                [descricao, codigo, quantidade, data_venda.strftime('%Y-%m-%d')]
                            )
                            print(
                                f"[OK] Peça encontrada: {descricao} | Data: {data_venda.strftime('%Y-%m-%d')} | Quantidade: {quantidade}")
                except Exception as e:
                    print(f"[ERRO] Erro na linha {index}: {e}")

        except Exception as e:
            print(f"[ERRO] Falha ao ler .xlsx: {e}")
            return

        if pecas_compradas_ano:
            print(f"[RESULTADO] Total de peças compradas no último ano: {len(pecas_compradas_ano)}")
            exibir_pecas_compradas_ano(pecas_compradas_ano)
        else:
            print(f"[AVISO] Nenhuma peça foi comprada por {cliente} no último ano.")
            messagebox.showinfo("Aviso", f"Nenhuma peça foi comprada por '{cliente}' no último ano.")

        # === Busca informações da máquina ===
        caminho_clientes = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\planilhas\clientes.xlsx"
        print("[INFO] Verificando dados da máquina...")

        df_clientes = pd.read_excel(caminho_clientes, header=None)
        cliente_info = df_clientes[(df_clientes[0] == cliente) & (df_clientes[1] == maquina)]

        if not cliente_info.empty:
            print(f"[OK] Dados encontrados: {cliente_info.iloc[0].tolist()}")
        else:
            print(f"[ERRO] Nenhum dado encontrado para o cliente '{cliente}' e máquina '{maquina}'.")
            return

        pecas = []
        colunas = [2, 3, 4, 5]
        pastas = ["maquina", "fonte", "cnc", "acessorios"]

        for i, pasta in enumerate(pastas):
            if pd.notna(cliente_info.iloc[0, colunas[i]]):
                nome_arquivo = str(cliente_info.iloc[0, colunas[i]]) + ".xlsx"
                caminho_arquivo = os.path.join(
                    r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\planilhas\pecas",
                    pasta, nome_arquivo
                )

                print(f"[BUSCA] Procurando arquivo: {caminho_arquivo}")

                if os.path.exists(caminho_arquivo):
                    df_pecas = pd.read_excel(caminho_arquivo, header=None)
                    pecas.extend(df_pecas.iloc[1:].values.tolist())
                    print(f"[OK] Peças carregadas da pasta '{pasta}'")
                else:
                    print(f"[AVISO] Arquivo não encontrado: {caminho_arquivo}")

        print(f"[RESULTADO] Total de peças da máquina '{maquina}': {len(pecas)}")
        abrir_tela_selecao_pecas(f"{cliente} - {maquina}", pecas)

    def exibir_pecas_compradas_ano(pecas):
        print("[TELA] Abrindo janela com peças compradas no último ano...")

        tela = tk.Toplevel()
        tela.title("Peças Compradas no Último Ano")
        tela.geometry("850x400")
        tela.grab_set()

        frame = ttk.Frame(tela)
        frame.pack(fill=tk.BOTH, expand=True)

        canvas = tk.Canvas(frame)
        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Cabeçalho
        headers = ["Data", "Código", "Descrição", "Quantidade"]
        for col, header in enumerate(headers):
            lbl = ttk.Label(scrollable_frame, text=header, font=("Arial", 12, "bold"), background="lightblue")
            lbl.grid(row=0, column=col, sticky="nsew", padx=5, pady=5)

        # Ajusta pesos das colunas para expandirem
        for col in range(len(headers)):
            scrollable_frame.grid_columnconfigure(col, weight=1)

        # Linhas de dados
        for i, peca in enumerate(pecas):
            row_color = 'lightgray' if i % 2 == 0 else 'white'

            ttk.Label(scrollable_frame, text=str(peca[3]), background=row_color, anchor="center").grid(row=i + 1,
                                                                                                       column=0,
                                                                                                       sticky="nsew",
                                                                                                       padx=5, pady=2)
            ttk.Label(scrollable_frame, text=str(peca[1]), background=row_color, anchor="w").grid(row=i + 1, column=1,
                                                                                                  sticky="nsew", padx=5,
                                                                                                  pady=2)
            ttk.Label(scrollable_frame, text=str(peca[0]), background=row_color, anchor="w").grid(row=i + 1, column=2,
                                                                                                  sticky="nsew", padx=5,
                                                                                                  pady=2)
            ttk.Label(scrollable_frame, text=str(peca[2]), background=row_color, anchor="e").grid(row=i + 1, column=3,
                                                                                                  sticky="nsew", padx=5,
                                                                                                  pady=2)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        ttk.Button(tela, text="Fechar", command=tela.destroy).pack(pady=10)



    clientes, df_clientes = carregar_clientes()

    # Widgets

    # Frame principal para organizar tudo
    frame_principal = ttk.Frame(tela_orcamento, padding=20)
    frame_principal.pack(fill=tk.BOTH, expand=True)

    # ==================== Linha do Cliente ====================
    frame_cliente = ttk.LabelFrame(frame_principal, text="Cliente", padding=(10, 5))
    frame_cliente.pack(fill=tk.X, pady=10)

    ttk.Label(frame_cliente, text="Nome do Cliente:", width=20, anchor="w").pack(side=tk.LEFT, padx=(0, 5))
    entrada_cliente = ttk.Entry(frame_cliente)
    entrada_cliente.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
    entrada_cliente.bind("<KeyRelease>", filtrar_clientes)

    combo_cliente = ttk.Combobox(frame_cliente, values=clientes)
    combo_cliente.pack(side=tk.LEFT, padx=(0, 5))
    combo_cliente.bind("<<ComboboxSelected>>", selecionar_cliente)

    # ==================== Linha da Máquina ====================
    frame_maquina = ttk.LabelFrame(frame_principal, text="Máquina", padding=(10, 5))
    frame_maquina.pack(fill=tk.X, pady=10)

    ttk.Label(frame_maquina, text="Nome da Máquina:", width=20, anchor="w").pack(side=tk.LEFT, padx=(0, 5))
    combo_maquina = ttk.Combobox(frame_maquina)
    combo_maquina.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
    combo_maquina.bind("<<ComboboxSelected>>", selecionar_maquina)

    # ==================== Caixa de filtro adicional ====================
    frame_filtros = ttk.LabelFrame(frame_principal, text="Filtros", padding=(10, 5))
    frame_filtros.pack(fill=tk.X, pady=10)

    filtro_estoque_var = tk.BooleanVar()
    ttk.Checkbutton(frame_filtros, text="Somente peças com estoque disponível", variable=filtro_estoque_var).pack(
        anchor="w")

# Função para consultar, cancelar ou gerar PDF de algum orçamento pendente
def consultar_orcamentos_pendentes():
    tela_consulta = tk.Toplevel()
    tela_consulta.title("Consultar Orçamentos")
    tela_consulta.geometry("600x400")
    tela_consulta.grab_set()

    ttk.Label(tela_consulta, text="Pesquisar Orçamento (ID ou Nome do Cliente):").pack()
    entrada_pesquisa = ttk.Entry(tela_consulta)
    entrada_pesquisa.pack()

    lista_orcamentos = tk.Listbox(tela_consulta)
    lista_orcamentos.pack(fill=tk.BOTH, expand=True)

    caminho_orcamentos = os.path.join("S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\orçamentos\\pendentes")
    arquivos = os.listdir(caminho_orcamentos)

    def atualizar_lista(event=None):
        termo = entrada_pesquisa.get().lower()
        lista_orcamentos.delete(0, tk.END)
        for arquivo in arquivos:
            if termo in arquivo.lower():
                lista_orcamentos.insert(tk.END, arquivo)

    entrada_pesquisa.bind("<KeyRelease>", atualizar_lista)

    def abrir_orcamento():
        selecao = lista_orcamentos.get(tk.ACTIVE)
        if selecao:
            caminho_arquivo = os.path.join(caminho_orcamentos, selecao)
            os.startfile(caminho_arquivo)  # Abrindo o Excel diretamente

    def cancelar_orcamento_manual():
        selecao = lista_orcamentos.get(tk.ACTIVE)
        if not selecao:
            messagebox.showwarning("Atenção", "Selecione um orçamento da lista.")
            return

        # Caminhos
        pasta_pendentes = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\orçamentos\pendentes"
        pasta_cancelados = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\orçamentos\cancelados"
        caminho_motivos = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\planilhas\motivos_de_cancelamento.txt"

        nome_arquivo = selecao
        caminho_arquivo = os.path.join(pasta_pendentes, nome_arquivo)

        if not os.path.exists(caminho_arquivo):
            messagebox.showerror("Erro", "Arquivo selecionado não encontrado.")
            return

        # Ler motivos
        try:
            with open(caminho_motivos, "r", encoding="utf-8") as f:
                motivos = [linha.strip() for linha in f.readlines() if linha.strip()]
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao ler o arquivo de motivos: {e}")
            return

        if not motivos:
            messagebox.showwarning("Aviso", "Arquivo de motivos está vazio.")
            return

        # Criar janela de cancelamento
        janela = tk.Toplevel()
        janela.title("Cancelar Orçamento")
        janela.geometry("500x500")
        janela.configure(bg="white")
        janela.grab_set()

        tk.Label(janela, text=f"Cancelar orçamento \"{nome_arquivo}\"?", bg="white", font=("Arial", 11),
                 wraplength=480).pack(pady=10)

        tk.Label(janela, text="Motivo do Cancelamento:", bg="white").pack()
        motivo_listbox = tk.Listbox(janela, height=6, width=60)
        for motivo in motivos:
            motivo_listbox.insert(tk.END, motivo)
        motivo_listbox.pack(pady=5)

        tk.Label(janela, text="Observações:", bg="white").pack()
        entrada_obs = tk.Text(janela, height=4, width=60)
        entrada_obs.pack(pady=5)

        def confirmar_cancelamento():
            selecionado = motivo_listbox.curselection()
            if not selecionado:
                messagebox.showwarning("Campo obrigatório", "Selecione um motivo.")
                return

            motivo_selecionado = motivo_listbox.get(selecionado[0])
            observacao = entrada_obs.get("1.0", tk.END).strip()

            try:
                wb = load_workbook(caminho_arquivo)
                ws = wb.active
                ws.cell(row=1, column=12).value = motivo_selecionado  # Coluna H (índice 7)
                ws.cell(row=1, column=13).value = observacao  # Coluna I (índice 8)
                wb.save(caminho_arquivo)
                wb.close()

                shutil.move(caminho_arquivo, os.path.join(pasta_cancelados, nome_arquivo))
                messagebox.showinfo("Sucesso", "Orçamento cancelado com sucesso.")
                janela.destroy()
                tela_consulta.destroy()  # Se quiser recarregar a lista após cancelamento

            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao salvar ou mover arquivo: {e}")

        frame_botoes = tk.Frame(janela, bg="white")
        frame_botoes.pack(pady=10)
        ttk.Button(frame_botoes, text="Confirmar", command=confirmar_cancelamento, width=15).pack(side="left", padx=10)
        ttk.Button(frame_botoes, text="Fechar", command=janela.destroy, width=15).pack(side="right", padx=10)

    def gerar_relatorio_pdf():
        selecao = lista_orcamentos.get(tk.ACTIVE)
        if not selecao:
            messagebox.showwarning("Atenção", "Nenhum orçamento selecionado!")
            return

        caminho_arquivo = os.path.join(caminho_orcamentos, selecao)
        if not os.path.exists(caminho_arquivo):
            messagebox.showerror("Erro", "O arquivo não foi encontrado.")
            return

        try:
            df = pd.read_excel(caminho_arquivo, header=None, skiprows=1)
            nome_cliente = str(df.iloc[0, 5])
            nome_pdf = selecao.replace(".xlsx", ".pdf")

            pasta_pdf_padrao = os.path.join(
                "S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\orçamentos\\pdfs"
            )
            if not os.path.exists(pasta_pdf_padrao):
                os.makedirs(pasta_pdf_padrao)
            caminho_pdf_padrao = os.path.join(pasta_pdf_padrao, nome_pdf)

            caminho_pdf_usuario = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("Arquivos PDF", "*.pdf")],
                initialfile=nome_pdf,
                title="Salvar relatório PDF como..."
            )
            if not caminho_pdf_usuario:
                return

            caminho_template_pdf = "S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\personalizacao\\template.pdf"  # Altere aqui
            logo_path = "S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\personalizacao\\logo.png"  # Altere aqui
            ALTURA_HEADER = 150
            ALTURA_FOOTER = 100
            MARGEM_SUPERIOR = 20
            MARGEM_INFERIOR = 20

            buffer_overlay = io.BytesIO()
            c_overlay = canvas.Canvas(buffer_overlay, pagesize=letter)
            page_width, page_height = letter

            itens_faltantes = []

            def quebrar_texto(texto, largura_maxima, fonte="Helvetica", tamanho=11):
                palavras = texto.split(" ")
                linhas = []
                linha_atual = ""
                for palavra in palavras:
                    linha_teste = f"{linha_atual} {palavra}".strip()
                    if c_overlay.stringWidth(linha_teste, fonte, tamanho) <= largura_maxima:
                        linha_atual = linha_teste
                    else:
                        if linha_atual:
                            linhas.append(linha_atual)
                        linha_atual = palavra
                if linha_atual:
                    linhas.append(linha_atual)
                return linhas

            def desenhar_cabecalho_overlay(canvas_overlay):
                canvas_overlay.setFont("Helvetica-Bold", 18)
                canvas_overlay.setFillColorRGB(0.2, 0.2, 0.8)
                canvas_overlay.drawString(50, 780, "Lista de peças preventiva")  # posição do título

                nome_cliente_quebrado = quebrar_texto(nome_cliente, 400)
                y_header = 760  # aumentar para cabeçalho subir
                for linha_cliente in nome_cliente_quebrado:
                    canvas_overlay.setFont("Helvetica", 12)
                    canvas_overlay.drawString(50, y_header, f"Cliente: {linha_cliente}")
                    y_header -= 20

                nome_maquina = str(df.iloc[0, 6]) if pd.notna(df.iloc[0, 6]) else "Não especificado"
                nome_maquina_quebrado = quebrar_texto(nome_maquina, 400)
                for linha_maquina in nome_maquina_quebrado:
                    canvas_overlay.setFont("Helvetica", 12)
                    canvas_overlay.drawString(50, y_header, f"Máquina: {linha_maquina}")
                    y_header -= 20

                canvas_overlay.setFont("Helvetica", 10)
                data_atual = datetime.now().strftime("%d/%m/%Y")
                canvas_overlay.drawString(50, y_header, f"Data: {data_atual} | Empresa: GRUPOBAW")

                if os.path.exists(logo_path):
                    canvas_overlay.drawImage(logo_path, 450, y_header + 10, width=100, height=50)

                y_header -= 40
                return y_header

            y_position = desenhar_cabecalho_overlay(c_overlay)

            for index, row in df.iterrows():
                if y_position < (ALTURA_FOOTER + MARGEM_INFERIOR + 200):
                    c_overlay.showPage()
                    y_position = desenhar_cabecalho_overlay(c_overlay)

                descricao = str(row[0]) if pd.notna(row[0]) else "Sem descrição"
                quantidade = str(row[1]) if pd.notna(row[1]) else "N/A"
                periodo = str(row[3]) if pd.notna(row[3]) else "N/A"
                if periodo != "N/A":
                    periodo = f"{int(float(periodo))} ano" if float(periodo) == 1.0 else f"{int(float(periodo))} anos"
                codigo = str(row[2]).split('.')[0] if pd.notna(row[2]) else "N/A"
                tempo_troca = str(row[4]) if pd.notna(row[4]) else "Não especificado"
                if tempo_troca != "Não especificado":
                    tempo_troca = f"{int(float(tempo_troca))} hora" if float(
                        tempo_troca) == 1 else f"{int(float(tempo_troca))} horas"

                pasta_item = os.path.join(
                    "S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\planilhas\\informacoes_detalhadas",
                    codigo
                )
                caminho_descricao = os.path.join(pasta_item, "descricao.txt")
                descricao_texto = "Descrição não disponível"

                if not os.path.exists(pasta_item):
                    itens_faltantes.append(codigo)
                else:
                    if os.path.exists(caminho_descricao):
                        with open(caminho_descricao, "r", encoding="utf-8") as f:
                            descricao_texto = f.read()

                imagem_localizacao = None
                imagem_peca = None
                for ext in ["jpg", "png"]:
                    caminho_img_loc = os.path.join(pasta_item, f"localizacao.{ext}")
                    caminho_img_peca = os.path.join(pasta_item, f"peca.{ext}")
                    if os.path.exists(caminho_img_loc):
                        imagem_localizacao = caminho_img_loc
                    if os.path.exists(caminho_img_peca):
                        imagem_peca = caminho_img_peca

                c_overlay.setFont("Helvetica-Bold", 14)
                c_overlay.setFillColorRGB(0.2, 0.6, 0.2)
                c_overlay.drawString(50, y_position, f"Peça: {descricao}")
                y_position -= 20

                c_overlay.setFont("Helvetica", 12)
                c_overlay.setFillColorRGB(0, 0, 0)
                c_overlay.drawString(50, y_position, f"Código: {codigo}")
                y_position -= 20
                c_overlay.drawString(50, y_position, f"Quantidade: {quantidade}")
                y_position -= 20
                c_overlay.drawString(50, y_position, f"Período de troca: {periodo}")
                y_position -= 20
                c_overlay.drawString(50, y_position, f"Tempo Previsto de Troca: {tempo_troca}")
                y_position -= 40

                texto_quebrado = quebrar_texto(descricao_texto, 500)
                c_overlay.setFont("Helvetica", 11)
                for linha in texto_quebrado:
                    c_overlay.drawString(50, y_position, linha)
                    y_position -= 15

                altura_img = 180
                largura_max = 500

                if imagem_localizacao:
                    try:
                        c_overlay.drawImage(ImageReader(imagem_localizacao), 50, y_position - altura_img,
                                            width=largura_max, height=altura_img, preserveAspectRatio=True)
                        c_overlay.drawString(50, y_position - altura_img - 15, "Localização da peça")
                        y_position -= altura_img + 30
                    except:
                        pass

                if imagem_peca:
                    try:
                        c_overlay.drawImage(ImageReader(imagem_peca), 50, y_position - altura_img, width=largura_max,
                                            height=altura_img, preserveAspectRatio=True)
                        c_overlay.drawString(50, y_position - altura_img - 15, "Imagem da peça")
                        y_position -= altura_img + 30
                    except:
                        pass

            c_overlay.save()
            buffer_overlay.seek(0)
            overlay_reader = PdfReader(buffer_overlay)
            writer = PdfWriter()

            for i in range(len(overlay_reader.pages)):
                template_reader = PdfReader(caminho_template_pdf)
                pagina_template = template_reader.pages[0]
                pagina_overlay = overlay_reader.pages[i]
                pagina_template.merge_page(pagina_overlay)
                writer.add_page(pagina_template)

            with open(caminho_pdf_usuario, "wb") as f_out:
                writer.write(f_out)

            if caminho_pdf_usuario != caminho_pdf_padrao:
                try:
                    with open(caminho_pdf_usuario, "rb") as f_src:
                        with open(caminho_pdf_padrao, "wb") as f_dest:
                            f_dest.write(f_src.read())
                except:
                    messagebox.showwarning("Atenção", "Não foi possível salvar automaticamente na pasta padrão.")

            if itens_faltantes:
                faltantes_str = "\n".join(itens_faltantes)
                messagebox.showwarning("Aviso",
                                       f"As seguintes peças estão sem informações detalhadas:\n{faltantes_str}")

            messagebox.showinfo("Sucesso", f"Relatório gerado com sucesso em:\n{caminho_pdf_usuario}")

        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao gerar PDF: {str(e)}")

    ttk.Button(tela_consulta, text="Abrir Orçamento", command=abrir_orcamento).pack()
    ttk.Button(tela_consulta, text="Cancelar", command=cancelar_orcamento_manual).pack()  # Alteração do botão para "Cancelar"
    ttk.Button(tela_consulta, text="Gerar PDF", command=gerar_relatorio_pdf).pack()  # Novo botão

    # Adicionando botão de "Voltar"
    ttk.Button(tela_consulta, text="Voltar", command=tela_consulta.destroy).pack(pady=10)

    atualizar_lista()

def consultar_orcamentos_cancelados():
    pasta_cancelados = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\orçamentos\cancelados"
    caminho_motivos = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\planilhas\motivos_de_cancelamento.txt"

    def carregar_dados():
        tree.delete(*tree.get_children())
        termo = entrada_pesquisa.get().lower()
        for arquivo in os.listdir(pasta_cancelados):
            if arquivo.endswith(".xlsx") and termo in arquivo.lower():
                caminho = os.path.join(pasta_cancelados, arquivo)
                try:
                    wb = load_workbook(caminho, data_only=True)
                    ws = wb.active
                    motivo = ws.cell(row=1, column=12).value or ""
                    obs = ws.cell(row=1, column=13).value or ""
                    wb.close()
                except:
                    motivo = obs = "Erro ao ler"
                tree.insert("", tk.END, values=(arquivo, motivo, obs))

    def abrir_arquivo():
        item = tree.focus()
        if not item:
            messagebox.showwarning("Seleção", "Selecione um orçamento.")
            return
        nome_arquivo = tree.item(item)['values'][0]
        caminho = os.path.join(pasta_cancelados, nome_arquivo)
        try:
            os.startfile(caminho)
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível abrir o arquivo:\n{e}")

    def apagar_arquivo():
        item = tree.focus()
        if not item:
            messagebox.showwarning("Seleção", "Selecione um orçamento.")
            return
        nome_arquivo = tree.item(item)['values'][0]
        caminho = os.path.join(pasta_cancelados, nome_arquivo)

        confirm = messagebox.askyesno("Confirmação", f"Tem certeza que deseja excluir \"{nome_arquivo}\"?")
        if confirm:
            try:
                os.remove(caminho)
                carregar_dados()
                messagebox.showinfo("Sucesso", "Arquivo apagado com sucesso.")
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao apagar:\n{e}")

    def editar_motivos():
        def carregar_motivos():
            listbox.delete(0, tk.END)
            try:
                with open(caminho_motivos, "r", encoding="utf-8") as f:
                    for linha in f:
                        listbox.insert(tk.END, linha.strip())
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao carregar motivos:\n{e}")

        def adicionar_motivo():
            motivo = entrada_motivo.get().strip()
            if motivo:
                listbox.insert(tk.END, motivo)
                entrada_motivo.delete(0, tk.END)

        def remover_selecionado():
            selecionado = listbox.curselection()
            for i in reversed(selecionado):
                listbox.delete(i)

        def salvar_motivos():
            try:
                with open(caminho_motivos, "w", encoding="utf-8") as f:
                    for i in range(listbox.size()):
                        f.write(listbox.get(i) + "\n")
                messagebox.showinfo("Sucesso", "Motivos salvos com sucesso.")
                janela_edicao.destroy()
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao salvar:\n{e}")

        # Janela de edição
        janela_edicao = tk.Toplevel(janela)
        janela_edicao.title("Editar Motivos de Cancelamento")
        janela_edicao.geometry("500x400")
        janela_edicao.configure(bg="white")
        janela_edicao.grab_set()

        listbox = tk.Listbox(janela_edicao, selectmode=tk.MULTIPLE, width=60, height=15)
        listbox.pack(pady=10)

        entrada_motivo = tk.Entry(janela_edicao, width=50)
        entrada_motivo.pack(pady=5)

        frame_botoes_motivos = tk.Frame(janela_edicao, bg="white")
        frame_botoes_motivos.pack(pady=5)

        ttk.Button(frame_botoes_motivos, text="Adicionar", command=adicionar_motivo).pack(side="left", padx=5)
        ttk.Button(frame_botoes_motivos, text="Remover Selecionado", command=remover_selecionado).pack(side="left", padx=5)
        ttk.Button(janela_edicao, text="Salvar Alterações", command=salvar_motivos).pack(pady=10)

        carregar_motivos()

    def reabrir_orcamento():
        item = tree.focus()
        if not item:
            messagebox.showwarning("Seleção", "Selecione um orçamento para reabrir.")
            return

        nome_arquivo = tree.item(item)['values'][0]
        caminho_origem = os.path.join(pasta_cancelados, nome_arquivo)
        pasta_pendentes = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\orçamentos\pendentes"
        caminho_destino = os.path.join(pasta_pendentes, nome_arquivo)

        try:
            # Limpar colunas 12 e 13
            wb = load_workbook(caminho_origem)
            ws = wb.active
            ws.cell(row=1, column=12).value = ""
            ws.cell(row=1, column=13).value = ""
            wb.save(caminho_origem)
            wb.close()

            # Mover o arquivo
            os.replace(caminho_origem, caminho_destino)

            carregar_dados()
            messagebox.showinfo("Sucesso", f"Orçamento \"{nome_arquivo}\" foi reaberto com sucesso e movido para os pendentes.")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao reabrir orçamento:\n{e}")

    # Janela principal
    janela = tk.Toplevel()
    janela.title("Orçamentos Cancelados")
    janela.geometry("900x500")
    janela.configure(bg="white")
    janela.grab_set()

    # Campo de pesquisa
    tk.Label(janela, text="Pesquisar Orçamento:", bg="white").pack(pady=5)
    entrada_pesquisa = tk.Entry(janela, width=50)
    entrada_pesquisa.pack()
    entrada_pesquisa.bind("<KeyRelease>", lambda e: carregar_dados())

    # Treeview
    colunas = ("arquivo", "motivo", "observacao")
    tree = ttk.Treeview(janela, columns=colunas, show="headings", height=15)
    tree.heading("arquivo", text="Nome do Arquivo")
    tree.heading("motivo", text="Motivo do Cancelamento")
    tree.heading("observacao", text="Observação")
    tree.column("arquivo", width=300)
    tree.column("motivo", width=250)
    tree.column("observacao", width=300)
    tree.pack(pady=10, fill=tk.BOTH, expand=True)

    # Botões
    frame_botoes = tk.Frame(janela, bg="white")
    frame_botoes.pack(pady=5)
    ttk.Button(frame_botoes, text="Abrir", command=abrir_arquivo, width=15).pack(side="left", padx=10)
    ttk.Button(frame_botoes, text="Apagar", command=apagar_arquivo, width=15).pack(side="left", padx=10)
    ttk.Button(frame_botoes, text="Editar motivos de cancelamento", command=editar_motivos, width=30).pack(side="left", padx=10)
    ttk.Button(frame_botoes, text="Reabrir Orçamento", command=reabrir_orcamento, width=20).pack(side="left", padx=10)

    carregar_dados()

# Função que verifica a validade dos orçamentos pendente e sugere o cancelamento dele, com a opção de selecionar uma motivação
def mover_orcamentos_antigos():
    from datetime import datetime, timedelta
    import tkinter as tk
    from tkinter import messagebox, ttk
    from openpyxl import load_workbook
    import os
    import shutil

    pasta_pendentes = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\orçamentos\pendentes"
    pasta_cancelados = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\orçamentos\cancelados"
    caminho_motivos = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\planilhas\motivos_de_cancelamento.txt"
    caminho_opcionais ="S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\personalizacao\\opcionais.txt"

    # Valor padrão de dias
    dias_limite = 0  # 0 dias = orçamentos criados hoje ainda são válidos

    try:
        with open(caminho_opcionais, "r") as f:
            linhas = [linha.strip() for linha in f.readlines()]
            if len(linhas) >= 2 and linhas[1].isdigit():
                dias_limite = int(linhas[1])
        print(f"[CONFIG] Dias para expiração dos orçamentos definidos como: {dias_limite}")
    except Exception as e:
        print(f"[AVISO] Não foi possível ler o arquivo 'opcionais.txt'. Usando valor padrão (0 dias). Erro: {e}")

    limite_tempo = datetime.now() - timedelta(days=dias_limite)

    # Tente abrir e ler os motivos de cancelamento
    try:
        with open(caminho_motivos, "r", encoding="utf-8") as f:
            motivos = [linha.strip() for linha in f.readlines() if linha.strip()]
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao ler o arquivo de motivos: {e}")
        return

    if not motivos:
        messagebox.showwarning("Aviso", "Nenhum motivo de cancelamento encontrado no arquivo.")
        return

    arquivos = [f for f in os.listdir(pasta_pendentes) if f.endswith(".xlsx")]

    def processar_proximo():
        if not arquivos:
            return

        nome_arquivo = arquivos.pop(0)
        caminho_arquivo = os.path.join(pasta_pendentes, nome_arquivo)
        data_criacao = datetime.fromtimestamp(os.path.getctime(caminho_arquivo))

        if data_criacao >= limite_tempo:
            processar_proximo()
            return

        janela = tk.Tk()
        janela.title("Orçamento Expirado")
        janela.geometry("500x500")
        janela.configure(bg="white")
        janela.grab_set()

        tk.Label(janela, text=f"O orçamento \"{nome_arquivo}\" expirou.\nDeseja movê-lo para a pasta de cancelados?",
                 bg="white", font=("Arial", 11), wraplength=480).pack(pady=10)

        tk.Label(janela, text="Motivo do Cancelamento:", bg="white").pack()

        motivo_var = tk.StringVar()
        motivo_listbox = tk.Listbox(janela, height=6, width=60)
        for motivo in motivos:
            motivo_listbox.insert(tk.END, motivo)
        motivo_listbox.pack(pady=5)

        tk.Label(janela, text="Observação:", bg="white").pack()
        entrada_obs = tk.Text(janela, height=4, width=60)
        entrada_obs.pack(pady=5)

        def confirmar():
            try:
                selecionado = motivo_listbox.curselection()
                if not selecionado:
                    messagebox.showwarning("Campo obrigatório", "Selecione um motivo de cancelamento.")
                    return

                motivo_selecionado = motivo_listbox.get(selecionado[0])
                obs = entrada_obs.get("1.0", "end").strip()

                wb = load_workbook(caminho_arquivo)
                ws = wb.active
                ws.cell(row=1, column=12).value = motivo_selecionado
                ws.cell(row=1, column=13).value = obs
                wb.save(caminho_arquivo)
                wb.close()

                shutil.move(caminho_arquivo, os.path.join(pasta_cancelados, nome_arquivo))
                print(f"Movido: {nome_arquivo}")
                messagebox.showinfo("Sucesso", f"O orçamento \"{nome_arquivo}\" foi cancelado com sucesso!")

            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao mover ou salvar: {e}")

            janela.destroy()
            processar_proximo()

        def adiar():
            janela.destroy()
            processar_proximo()

        frame_botoes = tk.Frame(janela, bg="white")
        frame_botoes.pack(pady=10)
        ttk.Button(frame_botoes, text="OK", command=confirmar, width=15).pack(side="left", padx=10)
        ttk.Button(frame_botoes, text="Adiar", command=adiar, width=15).pack(side="right", padx=10)

        janela.mainloop()

    processar_proximo()

def cadastro_preventivas():
        tela_consulta = tk.Toplevel()
        tela_consulta.title("Consultar Orçamentos")
        tela_consulta.geometry("600x800")
        tela_consulta.grab_set()

        ttk.Label(tela_consulta, text="Pesquisar Orçamento (ID ou Nome do Cliente):").pack()
        entrada_pesquisa = ttk.Entry(tela_consulta)
        entrada_pesquisa.pack()

        lista_orcamentos = tk.Listbox(tela_consulta)
        lista_orcamentos.pack(fill=tk.BOTH, expand=True)

        caminho_orcamentos = os.path.join("S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\orçamentos\\pendentes")
        arquivos = os.listdir(caminho_orcamentos)

        def atualizar_lista(event=None):
            termo = entrada_pesquisa.get().lower()
            lista_orcamentos.delete(0, tk.END)
            for arquivo in arquivos:
                if termo in arquivo.lower():
                    lista_orcamentos.insert(tk.END, arquivo)

        entrada_pesquisa.bind("<KeyRelease>", atualizar_lista)

        def confirmar_preventiva(orcamento):
            caminho_orcamento = os.path.join(
                "S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\orçamentos\\pendentes",
                orcamento)
            df_orcamento = pd.read_excel(caminho_orcamento)

            tela_pecas = tk.Toplevel()
            tela_pecas.title("Peças do Orçamento")
            tela_pecas.geometry("700x600")
            tela_pecas.grab_set()

            # Frame com altura fixa para não invadir os campos abaixo
            frame_lista = ttk.Frame(tela_pecas)
            frame_lista.pack(fill=tk.X, padx=10, pady=10)

            canvas = tk.Canvas(frame_lista, height=200)
            scrollbar = ttk.Scrollbar(frame_lista, orient=tk.VERTICAL, command=canvas.yview)
            scrollable_frame = ttk.Frame(canvas)

            scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)

            canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            pecas = df_orcamento.iloc[:, 0:5].values.tolist()
            for i, peca in enumerate(pecas):
                row_color = 'lightgray' if i % 2 == 0 else 'white'
                ttk.Label(scrollable_frame, text=f"{peca[0]}", background=row_color).grid(row=i, column=0, sticky="w",
                                                                                          padx=10, pady=5)
                ttk.Label(scrollable_frame, text=f"Qtd: {peca[1]}", background=row_color).grid(row=i, column=1,
                                                                                               sticky="w", padx=10,
                                                                                               pady=5)
                ttk.Label(scrollable_frame, text=f"Código: {peca[2]}", background=row_color).grid(row=i, column=2,
                                                                                                  sticky="w", padx=10,
                                                                                                  pady=5)
                ttk.Label(scrollable_frame, text=f"Período: {peca[3]}", background=row_color).grid(row=i, column=3,
                                                                                                   sticky="w", padx=10,
                                                                                                   pady=5)
                ttk.Label(scrollable_frame, text=f"MO: {peca[4]}", background=row_color).grid(row=i, column=4,
                                                                                              sticky="w", padx=10,
                                                                                              pady=5)

            # Função para validar os valores com ponto
            def validar_valor(entry):
                def on_validate(P):
                    P = P.replace(",", ".")
                    if P == "" or P.replace(".", "", 1).isdigit():
                        entry.delete(0, tk.END)
                        entry.insert(0, P)
                        return True
                    return False

                return (tela_pecas.register(on_validate), '%P')

            hoje = datetime.now().strftime("%d/%m/%Y")

            ttk.Label(tela_pecas, text="Data de Início (Formato: DD/MM/AAAA):").pack(pady=5)
            entrada_data_inicio = ttk.Entry(tela_pecas)
            entrada_data_inicio.insert(0, hoje)
            entrada_data_inicio.pack(pady=5)

            ttk.Label(tela_pecas, text="Data de Fim (Formato: DD/MM/AAAA):").pack(pady=5)
            entrada_data_fim = ttk.Entry(tela_pecas)
            entrada_data_fim.insert(0, hoje)
            entrada_data_fim.pack(pady=5)

            ttk.Label(tela_pecas, text="Valor Total das Peças:").pack(pady=5)
            entrada_valor_pecas = ttk.Entry(tela_pecas, validate="key")
            entrada_valor_pecas['validatecommand'] = validar_valor(entrada_valor_pecas)
            entrada_valor_pecas.pack(pady=5)

            ttk.Label(tela_pecas, text="Valor Total da Mão de Obra:").pack(pady=5)
            entrada_valor_mo = ttk.Entry(tela_pecas, validate="key")
            entrada_valor_mo['validatecommand'] = validar_valor(entrada_valor_mo)
            entrada_valor_mo.pack(pady=5)

            def salvar_preventiva():
                data_inicio = entrada_data_inicio.get()
                data_fim = entrada_data_fim.get()
                valor_pecas = entrada_valor_pecas.get().replace(",", ".")
                valor_mo = entrada_valor_mo.get().replace(",", ".")

                try:
                    datetime.strptime(data_inicio, '%d/%m/%Y')
                    datetime.strptime(data_fim, '%d/%m/%Y')
                except ValueError:
                    messagebox.showerror("Erro", "As datas inseridas estão em formato inválido. Use DD/MM/AAAA.")
                    return

                df_orcamento.at[0, 'Data Início'] = data_inicio
                df_orcamento.at[0, 'Data Fim'] = data_fim
                df_orcamento.at[0, 'Valor das Peças'] = valor_pecas
                df_orcamento.at[0, 'Valor Mão de Obra'] = valor_mo

                caminho_confirmados = os.path.join(
                    "S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\orçamentos\\confirmados")
                os.makedirs(caminho_confirmados, exist_ok=True)

                nome_arquivo = f"Orçamento Confirmado - {orcamento}"
                caminho_arquivo = os.path.join(caminho_confirmados, nome_arquivo)

                df_orcamento.columns = ['Peça', 'Qtd', 'Código', 'Período', 'MO', 'cliente', 'Máquina', 'Status Atualização', 'Data Início',
                                        'Data Fim', 'Valor das Peças', 'Valor Mão de Obra', 'Saldo total','Saldo disponível']
                df_orcamento.to_excel(caminho_arquivo, index=False)

                # 🟨 Mover o orçamento original da pasta "pendentes" para "historico_orçamentos"
                caminho_historico = os.path.join(
                    "S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\orçamentos\\historico_orçamentos")
                os.makedirs(caminho_historico, exist_ok=True)

                caminho_orcamento_antigo = os.path.join(
                    "S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\orçamentos\\pendentes",
                    orcamento)
                destino_orcamento_antigo = os.path.join(caminho_historico, orcamento)

                shutil.move(caminho_orcamento_antigo, destino_orcamento_antigo)

                messagebox.showinfo("Sucesso", "Preventiva confirmada com sucesso!")
                tela_pecas.destroy()
                tela_consulta.destroy()

            ttk.Button(tela_pecas, text="Confirmar Preventiva", command=salvar_preventiva).pack(pady=10)
            ttk.Button(tela_pecas, text="Voltar", command=tela_pecas.destroy).pack(pady=10)

        def criar_lista_orcamentos():
            # Limpar a lista de orçamentos antes de atualizar
            for widget in tela_consulta.winfo_children():
                widget.destroy()

            ttk.Label(tela_consulta, text="Pesquisar Orçamento (ID ou Nome do Cliente):").pack()
            entrada_pesquisa = ttk.Entry(tela_consulta)
            entrada_pesquisa.pack()

            caminho_orcamentos = os.path.join("S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\orçamentos\\pendentes")
            arquivos = os.listdir(caminho_orcamentos)

            def atualizar_lista(event=None):
                termo = entrada_pesquisa.get().lower()
                lista_orcamentos.delete(0, tk.END)
                for arquivo in arquivos:
                    if termo in arquivo.lower():
                        lista_orcamentos.insert(tk.END, arquivo)

            entrada_pesquisa.bind("<KeyRelease>", atualizar_lista)

            lista_orcamentos = tk.Listbox(tela_consulta, height=20)
            lista_orcamentos.pack(fill=tk.BOTH, expand=True)

            # Adicionar o botão de confirmar apenas uma vez por orçamento
            for arquivo in arquivos:
                lista_orcamentos.insert(tk.END, arquivo)

            def confirmar_preventiva_lista():
                orcamento_selecionado = lista_orcamentos.get(lista_orcamentos.curselection())
                confirmar_preventiva(orcamento_selecionado)

            ttk.Button(tela_consulta, text="Confirmar Preventiva", command=confirmar_preventiva_lista).pack(pady=5)

            atualizar_lista()

        criar_lista_orcamentos()

        # Botões de navegação
        ttk.Button(tela_consulta, text="Voltar", command=tela_consulta.destroy).pack(pady=10)

# Função para abrir a tela de peças avulsas  ** função descontinuada
def abrir_tela_pecas_avulsas():
    tela_preventivas = tk.Toplevel()
    tela_preventivas.title("Lançamento de Peças Avulsas")
    tela_preventivas.geometry("800x700")
    tela_preventivas.grab_set()

    # Carregar clientes do arquivo de clientes.xlsx
    def carregar_clientes():
        clientes = []
        path = "S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\planilhas\\clientes.xlsx"
        if os.path.exists(path):
            wb = openpyxl.load_workbook(path)
            sheet = wb.active
            for row in range(2, sheet.max_row + 1):  # Ignora o cabeçalho
                cliente = sheet.cell(row=row, column=1).value
                if cliente:
                    clientes.append(cliente.strip())  # Garante que não tenha espaços extras
        return clientes

    # Carregar nomes dos clientes com peças avulsas lançadas
    def carregar_clientes_pecas():
        pasta_pecas = "S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\planilhas\\pecas avulsas"
        if not os.path.exists(pasta_pecas):
            return []
        return [f.replace(".xlsx", "").strip() for f in os.listdir(pasta_pecas) if f.endswith(".xlsx")]

    # Pesquisa dinâmica para o campo de cliente (lançamento)
    def pesquisar_cliente(event, entry_cliente, clientes, listbox_clientes):
        query = entry_cliente.get().lower()
        listbox_clientes.delete(0, tk.END)
        if query:
            for cliente in clientes:
                if query in cliente.lower():
                    listbox_clientes.insert(tk.END, cliente)

    # Pesquisa dinâmica para o campo de cliente (consulta)
    def pesquisar_cliente_consulta(event, entry_consultar_cliente, clientes, listbox_clientes_consulta):
        query = entry_consultar_cliente.get().lower()
        listbox_clientes_consulta.delete(0, tk.END)
        if query:
            for cliente in clientes:
                if query in cliente.lower():
                    listbox_clientes_consulta.insert(tk.END, cliente)

    # Alteração para preencher o campo de pesquisa com o cliente selecionado
    def preencher_cliente_selecionado(event, entry_cliente, listbox_clientes):
        entry_cliente.delete(0, tk.END)  # Limpa o campo
        entry_cliente.insert(0, listbox_clientes.get(tk.ACTIVE))  # Preenche com o nome do cliente selecionado

    def preencher_cliente_selecionado_consulta(event, entry_consultar_cliente, listbox_clientes_consulta):
        entry_consultar_cliente.delete(0, tk.END)  # Limpa o campo
        entry_consultar_cliente.insert(0, listbox_clientes_consulta.get(tk.ACTIVE))  # Preenche com o nome do cliente selecionado

    # Salvar peça para um cliente
    def salvar_peca(entry_cliente, entry_descricao_peca, entry_codigo_peca, entry_quantidade, entry_data_venda):
        cliente = entry_cliente.get().strip()
        descricao_peca = entry_descricao_peca.get().strip()
        codigo_peca = entry_codigo_peca.get().strip()
        quantidade = entry_quantidade.get().strip()
        data_venda = entry_data_venda.get().strip()

        if not cliente or not descricao_peca or not codigo_peca or not quantidade or not data_venda:
            messagebox.showerror("Erro", "Todos os campos são obrigatórios.")
            return

        try:
            datetime.strptime(data_venda, "%d/%m/%Y")
        except ValueError:
            messagebox.showerror("Erro", "Data inválida. Use o formato DD/MM/AAAA.")
            return

        pasta_destino = "S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\planilhas\\pecas avulsas"
        os.makedirs(pasta_destino, exist_ok=True)

        nome_arquivo = f"{cliente}.xlsx"  # Mantém espaços no nome do cliente
        path = os.path.join(pasta_destino, nome_arquivo)

        if os.path.exists(path):
            wb = openpyxl.load_workbook(path)
            sheet = wb.active
        else:
            wb = Workbook()
            sheet = wb.active
            sheet.append(["Descrição", "Código", "Quantidade", "Data de Venda"])

        sheet.append([descricao_peca, codigo_peca, quantidade, data_venda])
        wb.save(path)



    # Consultar peças de um cliente
    def consultar_pecas(entry_consultar_cliente, tree):
        cliente = entry_consultar_cliente.get().strip()
        pasta_pecas = "S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\planilhas\\pecas avulsas"

        if not os.path.exists(pasta_pecas):
            messagebox.showerror("Erro", "Pasta de peças avulsas não encontrada.")
            return

        arquivos_disponiveis = os.listdir(pasta_pecas)
        print("Arquivos encontrados:", arquivos_disponiveis)  # Debugging

        arquivos = {f.replace(".xlsx", "").strip().lower(): f for f in arquivos_disponiveis if f.endswith(".xlsx")}
        print("Mapeamento de arquivos:", arquivos)  # Debugging

        if cliente.lower() not in arquivos:
            messagebox.showerror("Erro", f"Arquivo do cliente '{cliente}' não encontrado.")
            return

        path = os.path.join(pasta_pecas, arquivos[cliente.lower()])
        print("Abrindo arquivo:", path)  # Debugging

        wb = openpyxl.load_workbook(path)
        sheet = wb.active

        for row in tree.get_children():
            tree.delete(row)

        for row in sheet.iter_rows(min_row=2, values_only=True):  # Pula o cabeçalho
            tree.insert("", "end", values=row)



    clientes = carregar_clientes()
    clientes_pecas = carregar_clientes_pecas()

    # Lançamento de peças
    tk.Label(tela_preventivas, text="Cliente").grid(row=0, column=0)
    entry_cliente = tk.Entry(tela_preventivas)
    entry_cliente.grid(row=0, column=1)

    listbox_clientes = tk.Listbox(tela_preventivas, height=5, width=30)
    listbox_clientes.grid(row=1, column=1)
    listbox_clientes.bind("<ButtonRelease-1>", lambda e: preencher_cliente_selecionado(e, entry_cliente, listbox_clientes))

    entry_cliente.bind("<KeyRelease>", lambda e: pesquisar_cliente(e, entry_cliente, clientes, listbox_clientes))

    tk.Label(tela_preventivas, text="Descrição da Peça").grid(row=2, column=0)
    entry_descricao_peca = tk.Entry(tela_preventivas)
    entry_descricao_peca.grid(row=2, column=1)

    tk.Label(tela_preventivas, text="Código da Peça").grid(row=3, column=0)
    entry_codigo_peca = tk.Entry(tela_preventivas)
    entry_codigo_peca.grid(row=3, column=1)

    tk.Label(tela_preventivas, text="Quantidade").grid(row=4, column=0)
    entry_quantidade = tk.Entry(tela_preventivas)
    entry_quantidade.grid(row=4, column=1)

    tk.Label(tela_preventivas, text="Data de Venda (DD/MM/AAAA)").grid(row=5, column=0)
    entry_data_venda = tk.Entry(tela_preventivas)
    entry_data_venda.grid(row=5, column=1)

    tk.Button(tela_preventivas, text="Lançar Peça", command=lambda: salvar_peca(entry_cliente, entry_descricao_peca, entry_codigo_peca, entry_quantidade, entry_data_venda)).grid(row=6, column=0, columnspan=2, pady=10)

    # Consulta de peças
    tk.Label(tela_preventivas, text="Consultar Cliente").grid(row=7, column=0)
    entry_consultar_cliente = tk.Entry(tela_preventivas)
    entry_consultar_cliente.grid(row=7, column=1)

    listbox_clientes_consulta = tk.Listbox(tela_preventivas, height=5, width=30)
    listbox_clientes_consulta.grid(row=8, column=1)
    listbox_clientes_consulta.bind("<ButtonRelease-1>", lambda e: preencher_cliente_selecionado_consulta(e, entry_consultar_cliente, listbox_clientes_consulta))

    entry_consultar_cliente.bind("<KeyRelease>", lambda e: pesquisar_cliente_consulta(e, entry_consultar_cliente, clientes_pecas, listbox_clientes_consulta))

    tree = ttk.Treeview(tela_preventivas, columns=("Descrição", "Código", "Quantidade", "Data de Venda"), show="headings")
    tree.grid(row=9, column=0, columnspan=2)

    for col in ("Descrição", "Código", "Quantidade", "Data de Venda"):
        tree.heading(col, text=col)

    tk.Button(tela_preventivas, text="Consultar Peças", command=lambda: consultar_pecas(entry_consultar_cliente, tree)).grid(row=10, column=0, columnspan=2, pady=10)
    # Botão de Voltar
    tk.Button(tela_preventivas, text="Voltar", command=tela_preventivas.destroy).grid(row=11, column=0, columnspan=2, pady=10)

def carregar_preventivas_concluidas():
    pasta_concluidos = 'S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\orçamentos\\concluídos'
    if not os.path.exists(pasta_concluidos):
        print(f"A pasta {pasta_concluidos} não existe.")
        return []

    arquivos = [f for f in os.listdir(pasta_concluidos) if f.endswith('.xlsx')]
    preventivas = []

    for arquivo in arquivos:
        caminho_arquivo = os.path.join(pasta_concluidos, arquivo)
        try:
            wb = openpyxl.load_workbook(caminho_arquivo)
            planilha = wb.active

            nome_cliente = planilha.cell(row=2, column=6).value
            data_inicio = planilha.cell(row=2, column=8).value
            data_fim = planilha.cell(row=2, column=9).value
            codigo_peca = planilha.cell(row=2, column=3).value  # Coluna 3 para o código da peça

            if nome_cliente:
                preventivas.append({
                    "nome_cliente": nome_cliente,
                    "data_inicio": data_inicio,
                    "data_fim": data_fim,
                    "codigo_peca": codigo_peca,
                    "caminho_arquivo": caminho_arquivo
                })
        except Exception as e:
            print(f"Erro ao carregar {arquivo}: {e}")
    return preventivas


def exibir_detalhes_preventiva(preventiva):
    if not os.path.exists(preventiva['caminho_arquivo']):
        messagebox.showerror("Erro", f"Arquivo {preventiva['caminho_arquivo']} não encontrado.")
        return

    nova_tela = tk.Toplevel()
    nova_tela.title(f"Detalhes da Preventiva - {preventiva['nome_cliente']}")
    nova_tela.geometry("800x600")
    nova_tela.grab_set()

    wb = openpyxl.load_workbook(preventiva['caminho_arquivo'])
    planilha = wb.active

    ttk.Label(nova_tela, text="Peças:", font=("Arial", 12)).pack(pady=10)

    for row in range(2, planilha.max_row + 1):
        nome_peca = planilha.cell(row=row, column=1).value  # Coluna 1: Nome da peça
        codigo_peca = planilha.cell(row=row, column=3).value  # Coluna 2: Código da peça
        if nome_peca:
            ttk.Label(nova_tela, text=f"{nome_peca} (Código: {codigo_peca})", font=("Arial", 10)).pack()

    tempo_total = planilha.cell(row=planilha.max_row, column=5).value
    ttk.Label(nova_tela, text=f"Tempo Total Previsto: {tempo_total}", font=("Arial", 12)).pack(pady=10)

    valor_pecas = sum(float(planilha.cell(row=row, column=9).value or 0) for row in range(2, planilha.max_row + 1))
    valor_mao_obra = sum(float(planilha.cell(row=row, column=10).value or 0) for row in range(2, planilha.max_row + 1))

    ttk.Label(nova_tela, text=f"Valor Total das Peças: R${valor_pecas:.2f}", font=("Arial", 12)).pack(pady=5)
    ttk.Label(nova_tela, text=f"Valor Total de Mão de Obra: R${valor_mao_obra:.2f}", font=("Arial", 12)).pack(pady=5)


def aplicar_filtros(filtro_cliente, filtro_data_inicio, filtro_data_fim, filtro_codigo_peca):
    preventivas = carregar_preventivas_concluidas()

    # Filtro por cliente
    if filtro_cliente:
        preventivas = [p for p in preventivas if filtro_cliente.lower() in p['nome_cliente'].lower()]

    # Filtro por data de início
    if filtro_data_inicio:
        try:
            filtro_data_inicio = datetime.strptime(filtro_data_inicio, "%d/%m/%Y")
            preventivas = [p for p in preventivas if p['data_inicio'] and p['data_inicio'] >= filtro_data_inicio]
        except ValueError:
            messagebox.showerror("Erro", "Data de início inválida. Use o formato dd/mm/aaaa.")
            return []

    # Filtro por data de fim
    if filtro_data_fim:
        try:
            filtro_data_fim = datetime.strptime(filtro_data_fim, "%d/%m/%Y")
            preventivas = [p for p in preventivas if p['data_fim'] and p['data_fim'] <= filtro_data_fim]
        except ValueError:
            messagebox.showerror("Erro", "Data de fim inválida. Use o formato dd/mm/aaaa.")
            return []

    # Filtro por código de peça (agora percorrendo todas as linhas da coluna 3)
    if filtro_codigo_peca:
        preventivas = [
            p for p in preventivas if any(
                str(p['codigo_peca']).startswith(filtro_codigo_peca)  # Filtra códigos de peça que começam com a string do filtro
                for p in carregar_peças_do_arquivo(p['caminho_arquivo'])  # Função que vai buscar peças
            )
        ]

    return preventivas


def carregar_peças_do_arquivo(caminho_arquivo):
    wb = openpyxl.load_workbook(caminho_arquivo)
    planilha = wb.active
    peças = []

    for row in range(2, planilha.max_row + 1):  # Percorre todas as linhas de peças na planilha
        nome_peca = planilha.cell(row=row, column=1).value  # Coluna 1: Nome da peça
        codigo_peca = planilha.cell(row=row, column=3).value  # Coluna 3: Código da peça
        if nome_peca and codigo_peca:
            peças.append({'nome_peca': nome_peca, 'codigo_peca': codigo_peca})

    return peças


def atualizar_preventivas_concluidas(tree, filtro_cliente, filtro_data_inicio, filtro_data_fim, filtro_codigo_peca):
    for item in tree.get_children():
        tree.delete(item)

    preventivas = aplicar_filtros(filtro_cliente, filtro_data_inicio, filtro_data_fim, filtro_codigo_peca)

    if not preventivas:
        messagebox.showinfo("Sem Preventivas", "Não há preventivas que atendem aos critérios de filtro.")
        return

    for preventiva in preventivas:
        tree.insert("", "end", values=(
            preventiva['nome_cliente'],
            preventiva['data_inicio'],
            preventiva['data_fim'],
            preventiva['caminho_arquivo']))


def on_preventiva_select(event, tree):
    selected_item = tree.selection()
    if selected_item:
        caminho_arquivo = tree.item(selected_item[0])['values'][3]
        preventivas = carregar_preventivas_concluidas()
        for preventiva in preventivas:
            if preventiva['caminho_arquivo'] == caminho_arquivo:
                exibir_detalhes_preventiva(preventiva)
                break


def tela_consulta_preventivas():
    nova_tela = tk.Toplevel()
    nova_tela.title("Consulta Preventivas Concluídas")
    nova_tela.geometry("900x600")
    nova_tela.grab_set()

    # Filtros na parte superior da tela
    ttk.Label(nova_tela, text="Filtrar por Cliente:").grid(row=0, column=0, padx=10, pady=5)
    filtro_cliente = ttk.Entry(nova_tela)
    filtro_cliente.grid(row=0, column=1, padx=10, pady=5)

    ttk.Label(nova_tela, text="Filtrar por Data Início:").grid(row=0, column=2, padx=10, pady=5)
    filtro_data_inicio = ttk.Entry(nova_tela)
    filtro_data_inicio.grid(row=0, column=3, padx=10, pady=5)

    ttk.Label(nova_tela, text="Filtrar por Data Fim:").grid(row=0, column=4, padx=10, pady=5)
    filtro_data_fim = ttk.Entry(nova_tela)
    filtro_data_fim.grid(row=0, column=5, padx=10, pady=5)

    ttk.Label(nova_tela, text="Filtrar por Código da Peça:").grid(row=1, column=0, padx=10, pady=5)
    filtro_codigo_peca = ttk.Entry(nova_tela)
    filtro_codigo_peca.grid(row=1, column=1, padx=10, pady=5)

    # Botão para aplicar os filtros
    ttk.Button(nova_tela, text="Aplicar Filtros", command=lambda: atualizar_preventivas_concluidas(
        tree, filtro_cliente.get(), filtro_data_inicio.get(), filtro_data_fim.get(), filtro_codigo_peca.get())
    ).grid(row=1, column=2, padx=10, pady=5)

    # Criação da Treeview para exibir as preventivas
    tree = ttk.Treeview(nova_tela, columns=("Nome", "Início", "Fim", "Caminho"), show='headings')
    tree.heading("Nome", text="Cliente")
    tree.heading("Início", text="Data Início")
    tree.heading("Fim", text="Data Fim")
    tree.column("Caminho", width=0, stretch=tk.NO)  # Coluna oculta para armazenar o caminho do arquivo
    tree.grid(row=2, column=0, columnspan=6, pady=10, padx=10, sticky='nsew')

    # Atualizando a Treeview com os dados das preventivas
    atualizar_preventivas_concluidas(tree, "", "", "", "")

    # Adicionando funcionalidade de clique duplo para ver os detalhes
    tree.bind("<Double-1>", lambda e: on_preventiva_select(e, tree))

    nova_tela.mainloop()


def carregar_tickets():
    pasta_orcamentos = 'S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\orçamentos\\confirmados'
    if not os.path.exists(pasta_orcamentos):
        print(f"A pasta {pasta_orcamentos} não existe.")
        return []

    arquivos = [f for f in os.listdir(pasta_orcamentos) if f.endswith('.xlsx')]
    tickets = []

    for arquivo in arquivos:
        caminho_arquivo = os.path.join(pasta_orcamentos, arquivo)
        try:
            wb = openpyxl.load_workbook(caminho_arquivo)
            planilha = wb.active

            nome_cliente = planilha.cell(row=2, column=6).value
            data_inicio = planilha.cell(row=2, column=9).value
            data_fim = planilha.cell(row=2, column=10).value

            if nome_cliente:
                tickets.append({
                    "nome_cliente": nome_cliente,
                    "data_inicio": data_inicio,
                    "data_fim": data_fim,
                    "caminho_arquivo": caminho_arquivo
                })
        except Exception as e:
            print(f"Erro ao carregar {arquivo}: {e}")
    return tickets


def mover_arquivo(ticket, destino):
    pasta_destino = os.path.join("S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\orçamentos", destino)  # Caminho absoluto
    if not os.path.exists(pasta_destino):
        os.makedirs(pasta_destino)
    if os.path.exists(ticket['caminho_arquivo']):
        shutil.move(ticket['caminho_arquivo'], os.path.join(pasta_destino, os.path.basename(ticket['caminho_arquivo'])))


def exibir_detalhes(ticket):
    def formatar_data(data):
        if isinstance(data, datetime):
            return data.strftime("%d/%m/%Y")
        try:
            return datetime.strptime(str(data), "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y")
        except:
            try:
                return datetime.strptime(str(data), "%d/%m/%Y").strftime("%d/%m/%Y")
            except:
                return str(data)

    if not os.path.exists(ticket['caminho_arquivo']):
        print(f"Arquivo {ticket['caminho_arquivo']} não encontrado.")
        return

    nova_tela = tk.Toplevel()
    nova_tela.title(f"Detalhes do Orçamento - {ticket['nome_cliente']}")
    nova_tela.geometry("800x600")
    nova_tela.grab_set()

    wb = openpyxl.load_workbook(ticket['caminho_arquivo'])
    planilha = wb.active

    ttk.Label(nova_tela, text="Peças:", font=("Arial", 12)).pack(pady=10)

    for row in range(2, planilha.max_row + 1):
        peça = planilha.cell(row=row, column=1).value
        if peça:
            ttk.Label(nova_tela, text=peça, font=("Arial", 10)).pack()

    tempo_total = planilha.cell(row=planilha.max_row, column=5).value
    ttk.Label(nova_tela, text=f"Tempo Total Previsto: {tempo_total}", font=("Arial", 12)).pack(pady=10)

    valor_pecas = sum(float(planilha.cell(row=row, column=11).value or 0) for row in range(2, planilha.max_row + 1))
    valor_mao_obra = sum(float(planilha.cell(row=row, column=12).value or 0) for row in range(2, planilha.max_row + 1))

    ttk.Label(nova_tela, text=f"Valor Total das Peças: R${valor_pecas:.2f}", font=("Arial", 12)).pack(pady=5)
    ttk.Label(nova_tela, text=f"Valor Total de Mão de Obra: R${valor_mao_obra:.2f}", font=("Arial", 12)).pack(pady=5)

    # Campos para editar datas
    frame_datas = ttk.Frame(nova_tela)
    frame_datas.pack(pady=10)

    ttk.Label(frame_datas, text="Data Início (dd/mm/aaaa):").grid(row=0, column=0, padx=5, pady=5)
    entrada_inicio = ttk.Entry(frame_datas)
    entrada_inicio.insert(0, formatar_data(ticket['data_inicio']))
    entrada_inicio.grid(row=0, column=1, padx=5, pady=5)

    ttk.Label(frame_datas, text="Data Fim (dd/mm/aaaa):").grid(row=1, column=0, padx=5, pady=5)
    entrada_fim = ttk.Entry(frame_datas)
    entrada_fim.insert(0, formatar_data(ticket['data_fim']))
    entrada_fim.grid(row=1, column=1, padx=5, pady=5)

    def salvar_datas():
        try:
            data_inicio_str = entrada_inicio.get()
            data_fim_str = entrada_fim.get()

            cel_inicio = planilha.cell(row=2, column=9)
            cel_fim = planilha.cell(row=2, column=10)

            # Salva o texto puro, sem apóstrofo
            cel_inicio.value = data_inicio_str
            cel_fim.value = data_fim_str

            wb.save(ticket['caminho_arquivo'])
            messagebox.showinfo("Sucesso", "Datas atualizadas com sucesso!")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar datas: {e}")

    ttk.Button(nova_tela, text="Salvar Datas", command=salvar_datas).pack(pady=10)

    ttk.Button(nova_tela, text="Finalizado",
               command=lambda: [mover_arquivo(ticket, 'concluídos'), nova_tela.destroy()]).pack(pady=5)
    ttk.Button(nova_tela, text="Cancelado",
               command=lambda: [mover_arquivo(ticket, 'cancelados'), nova_tela.destroy()]).pack(pady=5)


def exibir_nome_ticket(event, calendar, tickets):
    # Obtém a data que está sendo selecionada no calendário
    data_selecionada = calendar.get_date()
    data_selecionada = datetime.strptime(data_selecionada, "%m/%d/%Y").date()  # Formato de data do calendário

    # Procurar o ticket correspondente à data selecionada
    for ticket in tickets:
        try:
            # Garantir que as datas de início e fim são datetime
            if isinstance(ticket['data_inicio'], str):
                data_inicio = datetime.strptime(ticket['data_inicio'], "%d/%m/%Y").date()
            else:
                data_inicio = ticket['data_inicio'].date()

            if isinstance(ticket['data_fim'], str):
                data_fim = datetime.strptime(ticket['data_fim'], "%d/%m/%Y").date()
            else:
                data_fim = ticket['data_fim'].date()

            # Verifica se a data selecionada está dentro do intervalo do ticket
            if data_inicio <= data_selecionada <= data_fim:
                # Exibe o nome do cliente associado ao ticket
                tooltip_text = f"Ticket de {ticket['nome_cliente']}"
                calendar.set_tooltip(event, tooltip_text)  # Atualiza o tooltip
                return
        except Exception as e:
            print(f"Erro ao verificar o ticket: {e}")

    # Se não encontrar nenhum ticket para a data, limpa o tooltip
    calendar.set_tooltip(event, "")


def atualizar_tickets(tree, calendar):
    for item in tree.get_children():
        tree.delete(item)
    tickets = carregar_tickets()
    datas_agendadas = []

    for ticket in tickets:
        tree.insert("", "end", values=(
            ticket['nome_cliente'], ticket['data_inicio'], ticket['data_fim'], ticket['caminho_arquivo']))

        if ticket['data_inicio'] and ticket['data_fim']:
            try:
                # Garantir que as datas estejam no formato datetime
                if isinstance(ticket['data_inicio'], str):
                    data_inicio = datetime.strptime(ticket['data_inicio'], "%d/%m/%Y")
                else:
                    data_inicio = ticket['data_inicio']

                if isinstance(ticket['data_fim'], str):
                    data_fim = datetime.strptime(ticket['data_fim'], "%d/%m/%Y")
                else:
                    data_fim = ticket['data_fim']

                # Adiciona todas as datas entre data_inicio e data_fim
                current_date = data_inicio
                while current_date <= data_fim:
                    datas_agendadas.append(current_date.date())  # Armazena só a data, sem a hora
                    current_date += timedelta(days=1)  # Avança um dia

            except ValueError:
                print(f"Formato de data inválido para o ticket {ticket['nome_cliente']}")
                continue

    calendar.calevent_remove('all')  # Remove eventos antigos
    for data in datas_agendadas:
        calendar.calevent_create(data, "Preventiva", 'reminder')

    # Configura o evento para exibir o nome do ticket ao passar o mouse
    calendar.bind("<<CalendarDayClicked>>", lambda event, calendar=calendar, tickets=tickets: exibir_nome_ticket(event, calendar, tickets))

    tree.after(5000, lambda: atualizar_tickets(tree, calendar))


def on_ticket_select(event, tree):
    selected_item = tree.selection()
    if selected_item:
        caminho_arquivo = tree.item(selected_item[0])['values'][3]
        tickets = carregar_tickets()
        for ticket in tickets:
            if ticket['caminho_arquivo'] == caminho_arquivo:
                exibir_detalhes(ticket)
                break


DIRETORIO_ORCAMENTOS = "S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\orçamentos\\concluídos"
DIRETORIO_NOTIFICACOES = "S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\notificacoes"  # Adiciona o diretório de notificações

# Função para garantir que o DataFrame tenha pelo menos 15 colunas
def garantir_colunas_suficientes(df):
    colunas_necessarias = 15
    if df.shape[1] < colunas_necessarias:
        for i in range(df.shape[1], colunas_necessarias):
            df[f"Coluna_{i + 1}"] = ""
    return df

# Função para atualizar a lista de notificações (depois de marcar como "verificado")
def atualizar_notificacoes(notificacoes_frame, root):
    verificar_notificacoes(notificacoes_frame, root)  # Recarregar as notificações

# Função para marcar como verificado na planilha
def marcar_como_verificado(arquivo, notificacoes_frame, root):
    try:
        df = pd.read_excel(arquivo, engine="openpyxl")

        df = garantir_colunas_suficientes(df)

        if df.iloc[0, 14].strip().lower() != "verificado":
            df.iloc[0, 14] = "verificado"

            df.to_excel(arquivo, index=False, engine="openpyxl")
            print(f"Arquivo {arquivo} marcado como verificado.")

            atualizar_notificacoes(notificacoes_frame, root)
        else:
            print(f"Arquivo {arquivo} já está verificado. Ignorando notificação.")

    except Exception as e:
        print(f"Erro ao marcar como verificado o arquivo {arquivo}: {e}")

    atualizar_notificacoes(notificacoes_frame, root)

# Função que verifica as notificações
def verificar_notificacoes(notificacoes_frame, root):
    import textwrap
    from datetime import datetime
    import os
    import pandas as pd
    from tkinter import ttk

    # Caminho do arquivo de configuração
    CAMINHO_OPCIONAIS = "S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\personalizacao\\opcionais.txt"

    # Valor padrão caso não consiga ler o arquivo
    dias_para_vencimento = 1

    # Tenta ler o número de dias do arquivo "opcionais.txt"
    try:
        with open(CAMINHO_OPCIONAIS, "r") as f:
            linhas = f.readlines()
            for linha in linhas:
                linha = linha.strip()
                if linha.isdigit():
                    dias_para_vencimento = int(linha)
                    break
        print(f"[CONFIG] Dias para vencimento definidos como: {dias_para_vencimento}")
    except Exception as e:
        print(f"[AVISO] Não foi possível ler o arquivo 'opcionais.txt'. Usando valor padrão: 1 dia. Erro: {e}")

    # Limpa notificações anteriores
    for widget in notificacoes_frame.winfo_children():
        widget.destroy()

    hoje = datetime.now().date()
    notificacao_gerada = False

    for diretorio in [DIRETORIO_ORCAMENTOS, DIRETORIO_NOTIFICACOES]:
        for arquivo in os.listdir(diretorio):
            if arquivo.endswith(".xlsx"):
                caminho_arquivo = os.path.join(diretorio, arquivo)

                try:
                    df = pd.read_excel(caminho_arquivo, engine="openpyxl")

                    while df.shape[0] < 2:
                        df.loc[df.shape[0]] = [""] * df.shape[1]

                    while df.shape[1] < 7:
                        df[f"Coluna_{df.shape[1] + 1}"] = ""

                    df.fillna("", inplace=True)

                    if df.shape[1] > 14 and df.iloc[0, 14].strip().lower() == "verificado":
                        continue

                    data_str = str(df.iloc[0, 8]).strip()

                    if not data_str:
                        continue

                    if data_str.startswith("'"):
                        data_str = data_str[1:]

                    try:
                        data_xlsx = pd.to_datetime(data_str, dayfirst=True, errors='coerce').date()

                        if data_xlsx is None or data_xlsx > hoje:
                            continue

                        if (hoje - data_xlsx).days > dias_para_vencimento:
                            nome_cliente = str(df.iloc[0, 5]).strip() if df.shape[1] > 5 else "Nome Desconhecido"
                            if not nome_cliente:
                                nome_cliente = "Nome Desconhecido"

                            codigo_peca = str(df.iloc[1, 0]).strip().split('.')[0]
                            texto_notificacao = f"Preventiva {nome_cliente} vencendo! - Data: {data_xlsx}"

                            wrapped_text = textwrap.fill(texto_notificacao, width=30)

                            botao_notificacao = ttk.Button(
                                notificacoes_frame,
                                text=wrapped_text,
                                command=lambda p=caminho_arquivo: abrir_detalhes_preventiva(p, root, notificacoes_frame),
                                style="Notificacao.TButton"
                            )

                            botao_notificacao.pack(pady=2, fill='x')
                            notificacao_gerada = True

                    except Exception as e:
                        print(f"Erro ao converter data no arquivo {arquivo}: {e}")

                except Exception as e:
                    print(f"Erro ao processar {arquivo}: {e}")

    if not notificacao_gerada:
        print("Nenhuma notificação foi gerada.")

def abrir_detalhes_preventiva(caminho_arquivo, root, notificacoes_frame):
    detalhes_janela = tk.Toplevel(root)
    detalhes_janela.title("Detalhes da Preventiva")
    detalhes_janela.geometry("400x400")
    detalhes_janela.grab_set()

    df = pd.read_excel(caminho_arquivo, engine="openpyxl")

    if len(df.columns) < 3:
        messagebox.showerror("Erro", "O arquivo não tem colunas suficientes para exibir peças.")
        return

    tree = ttk.Treeview(detalhes_janela, columns=("Peça", "Código"), show='headings')
    tree.heading("Peça", text="Peça")
    tree.heading("Código", text="Código")
    tree.pack(pady=10, fill=tk.BOTH, expand=True)

    for _, row in df.iterrows():
        peca = row.iloc[0] if len(row) > 0 else ""
        codigo = row.iloc[2] if len(row) > 2 else ""
        tree.insert("", tk.END, values=(peca, codigo))

    def marcar_verificado():
        if len(df.columns) > 14:
            df.iloc[:, 14] = "Verificado"
            df.to_excel(caminho_arquivo, index=False, engine="openpyxl")
            detalhes_janela.destroy()
            messagebox.showinfo("Sucesso", "Preventiva marcada como verificada!")
            atualizar_notificacoes(notificacoes_frame, root)
        else:
            messagebox.showerror("Erro", "O arquivo não tem a coluna necessária para marcar como verificado.")

    botao_verificar = ttk.Button(detalhes_janela, text="Verificado", command=marcar_verificado)
    botao_verificar.pack(pady=10)

# Função para abrir a tela de adicionar lembrete
def adicionar_lembrete():
    # Nova janela para adicionar o lembrete
    lembrete_janela = tk.Toplevel(root)
    lembrete_janela.title("Adicionar Lembrete")
    lembrete_janela.geometry("400x300")
    lembrete_janela.grab_set()

    # Campos para preencher
    tk.Label(lembrete_janela, text="Nome do Cliente").pack(pady=5)
    nome_cliente_entry = tk.Entry(lembrete_janela, width=30)
    nome_cliente_entry.pack(pady=5)

    tk.Label(lembrete_janela, text="Descrição do Lembrete").pack(pady=5)
    descricao_entry = tk.Entry(lembrete_janela, width=30)
    descricao_entry.pack(pady=5)

    tk.Label(lembrete_janela, text="Data do Lembrete").pack(pady=5)
    data_entry = tk.Entry(lembrete_janela, width=30)
    data_entry.insert(0, datetime.now().strftime("%d/%m/%Y"))  # Preenche com a data atual
    data_entry.pack(pady=5)

    # Função para salvar o lembrete
    def salvar_lembrete():
        nome_cliente = nome_cliente_entry.get()
        descricao = descricao_entry.get()
        data = data_entry.get()

        # Validação dos campos
        if not nome_cliente or not descricao or not data:
            messagebox.showerror("Erro", "Todos os campos devem ser preenchidos!")
            return

        try:
            # Cria um novo arquivo Excel
            caminho_arquivo = os.path.join("S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\notificacoes", f"{nome_cliente}.xlsx")
            if not os.path.exists("notificacoes"):
                os.makedirs("notificacoes")

            # Verifica se o arquivo já existe
            if os.path.exists(caminho_arquivo):
                df = pd.read_excel(caminho_arquivo, engine="openpyxl")
            else:
                # Se o arquivo não existir, cria um novo DataFrame com 15 colunas
                df = pd.DataFrame(columns=[f"Coluna_{i + 1}" for i in range(15)])

            # Adiciona a descrição, o nome do cliente e a data na linha
            nova_linha = [""] * 15
            nova_linha[0] = descricao  # Descrição na primeira coluna
            nova_linha[5] = nome_cliente  # Nome do cliente na coluna 5
            nova_linha[7] = f"'{data}"  # Data na coluna 8 (com apóstrofo)

            # Adiciona a nova linha ao DataFrame
            df.loc[len(df)] = nova_linha

            # Salva o arquivo Excel
            df.to_excel(caminho_arquivo, index=False, engine="openpyxl")

            messagebox.showinfo("Sucesso", "Lembrete adicionado com sucesso!")
            lembrete_janela.destroy()
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro ao salvar o lembrete: {e}")

    # Botão para salvar o lembrete
    ttk.Button(lembrete_janela, text="Salvar", command=salvar_lembrete).pack(pady=10)

def estatistica_geral():
    # Função para buscar os dados
    def obter_dados():
        dados = {}

        # Caminhos dos arquivos
        caminho_clientes = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\planilhas\clientes.xlsx"
        caminho_orcamentos_pendentes = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\orçamentos\pendentes"
        caminho_preventivas_confirmadas = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\orçamentos\confirmados"
        caminho_lembretes = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\notificacoes"
        caminho_pecas = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\planilhas\informacoes_detalhadas"

        # Clientes e máquinas
        try:
            df_clientes = pd.read_excel(caminho_clientes)
            dados["Total de clientes cadastrados"] = df_clientes.iloc[1:, 0].dropna().nunique()
            dados["Total de máquinas cadastradas"] = df_clientes.iloc[1:].shape[0]
        except:
            dados["Total de clientes cadastrados"] = "Erro"
            dados["Total de máquinas cadastradas"] = "Erro"

        # Orçamentos pendentes (contar apenas arquivos xlsx)
        try:
            arquivos = [f for f in os.listdir(caminho_orcamentos_pendentes) if f.endswith(".xlsx")]
            dados["Total de orçamentos gerados"] = len(arquivos)
        except:
            dados["Total de orçamentos gerados"] = "Erro"

        # Preventivas confirmadas (contar apenas arquivos xlsx)
        try:
            arquivos = [f for f in os.listdir(caminho_preventivas_confirmadas) if f.endswith(".xlsx")]
            dados["Total de preventivas cadastradas"] = len(arquivos)
        except:
            dados["Total de preventivas cadastradas"] = "Erro"

        # Lembretes ativos (contar linhas de todos os arquivos xlsx, exceto a primeira linha)
        try:
            arquivos = [f for f in os.listdir(caminho_lembretes) if f.endswith(".xlsx")]
            contador_ativos = 0
            for arquivo in arquivos:
                caminho_arquivo = os.path.join(caminho_lembretes, arquivo)
                try:
                    df = pd.read_excel(caminho_arquivo, header=None)
                    valor = str(df.iloc[1, 14]).strip().lower()  # linha 2 (índice 1), coluna 15 (índice 14)
                    if valor != "verificado":
                        contador_ativos += 1
                except:
                    continue  # Se der erro ao abrir o arquivo ou ler a célula, ignora esse arquivo
            dados["Total de lembretes ativos"] = contador_ativos
        except:
            dados["Total de lembretes ativos"] = "Erro"

        # Peças cadastradas (quantidade de pastas)
        try:
            total_pastas = len([f for f in os.listdir(caminho_pecas) if os.path.isdir(os.path.join(caminho_pecas, f))])
            dados["Total de peças cadastradas no sistema"] = total_pastas
        except:
            dados["Total de peças cadastradas no sistema"] = "Erro"

        return dados

    # Criação da interface gráfica
    janela = tk.Toplevel()
    janela.title("Estatísticas Gerais")
    janela.geometry("450x300")
    janela.configure(bg="#f4f4f4")
    janela.grab_set()

    dados = obter_dados()

    # Título
    titulo = tk.Label(janela, text="Estatísticas Gerais do Sistema", font=("Arial", 16, "bold"), bg="#f4f4f4")
    titulo.pack(pady=10)

    # Exibição dos dados
    frame = tk.Frame(janela, bg="#f4f4f4")
    frame.pack(pady=10)

    for chave, valor in dados.items():
        linha = tk.Frame(frame, bg="#f4f4f4")
        linha.pack(anchor="w", pady=2)
        lbl_nome = tk.Label(linha, text=f"{chave}:", font=("Arial", 12), bg="#f4f4f4")
        lbl_nome.pack(side="left")
        lbl_valor = tk.Label(linha, text=f" {valor}", font=("Arial", 12, "bold"), bg="#f4f4f4", fg="#333")
        lbl_valor.pack(side="left")

    # Botão de fechar
    btn_fechar = tk.Button(janela, text="Fechar", command=janela.destroy)
    btn_fechar.pack(pady=10)

    print("executado")

def estatisticas_comerciais():
    def carregar_dados():
        data_inicio = datetime.strptime(entry_inicio.get(), "%d/%m/%Y")
        data_fim = datetime.strptime(entry_fim.get(), "%d/%m/%Y")

        def arquivos_validos(pasta):
            arquivos = []
            for f in os.listdir(pasta):
                caminho = os.path.join(pasta, f)
                if f.endswith(".xlsx") and os.path.isfile(caminho):
                    data_criacao = datetime.fromtimestamp(os.path.getctime(caminho))
                    if data_inicio <= data_criacao <= data_fim:
                        arquivos.append(caminho)
            return arquivos

        confirmados = arquivos_validos(pasta_confirmados)
        concluidos = arquivos_validos(pasta_concluidos)
        pendentes = arquivos_validos(pasta_pendentes)
        cancelados = arquivos_validos(pasta_cancelados)

        convertidos = len(confirmados) + len(concluidos)
        total = convertidos + len(pendentes) + len(cancelados)
        taxa = (convertidos / total * 100) if total > 0 else 0

        ranking = Counter()
        lista_detalhada = []
        for arq in cancelados:
            try:
                wb = openpyxl.load_workbook(arq)
                ws = wb.active
                motivo = ws.cell(row=1, column=12).value or "Não informado"
                obs = ws.cell(row=1, column=9).value or ""
                ranking[motivo] += 1
                lista_detalhada.append((os.path.basename(arq), motivo, obs))
            except:
                continue

        atualizar_interface(convertidos, len(pendentes), len(cancelados), taxa, ranking, lista_detalhada, data_inicio, data_fim)

    def atualizar_interface(convertidos, pendentes, cancelados, taxa, ranking, lista_detalhada, data_inicio, data_fim):
        for widget in frame_info.winfo_children(): widget.destroy()
        for widget in frame_ranking.winfo_children(): widget.destroy()

        ttk.Label(frame_info, text=f"🟢 Orçamentos que viraram preventivas: {convertidos}").pack(anchor="w", pady=3)
        ttk.Label(frame_info, text=f"📊 Taxa de conversão: {taxa:.2f}%").pack(anchor="w", pady=3)
        ttk.Label(frame_info, text=f"🔴 Preventivas canceladas: {cancelados}").pack(anchor="w", pady=3)

        # Primeiro gráfico: distribuição geral
        fig.clear()
        ax1 = fig.add_subplot(121)
        ax1.pie([convertidos, pendentes, cancelados], labels=["Convertidos", "Pendentes", "Cancelados"],
                autopct='%1.1f%%', startangle=140, colors=["#4CAF50", "#FF9800", "#F44336"])
        ax1.set_title("Distribuição Comercial", fontsize=10)

        # Segundo gráfico: pizza de motivos de cancelamento
        ax2 = fig.add_subplot(122)
        if ranking:
            motivos = list(ranking.keys())
            qtds = list(ranking.values())
            ax2.pie(qtds, labels=motivos, autopct='%1.1f%%', startangle=140)
            ax2.set_title("Motivos de Cancelamento", fontsize=10)
        else:
            ax2.text(0.5, 0.5, "Sem dados de cancelamento", ha='center', va='center')
            ax2.axis('off')

        canvas.draw()

        ttk.Label(frame_ranking, text="🏆 Ranking de Cancelamentos:", font=("Arial", 10, "bold")).pack(anchor="w")
        for motivo, qtd in ranking.most_common():
            ttk.Label(frame_ranking, text=f"- {motivo}: {qtd} vez(es)").pack(anchor="w")

        frame_ranking.lista_detalhada = lista_detalhada
        frame_ranking.ranking = ranking
        frame_ranking.dados_resumo = (convertidos, pendentes, cancelados, taxa, data_inicio, data_fim)

    def exportar_pdf():
        caminho = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF Files", "*.pdf")])
        if not caminho:
            return

        convertidos, pendentes, cancelados, taxa, data_inicio, data_fim = frame_ranking.dados_resumo

        # Gerar imagem dos gráficos
        fig_path = "grafico_temp.png"
        fig.savefig(fig_path)

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)

        pdf.cell(0, 10, "Estatísticas Comerciais", ln=True, align='C')
        pdf.ln(5)
        pdf.set_font("Arial", size=11)
        pdf.cell(0, 10, f"Período analisado: {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}", ln=True)
        pdf.cell(0, 10, f"Orçamentos que viraram preventivas: {convertidos}", ln=True)
        pdf.cell(0, 10, f"Taxa de conversão: {taxa:.2f}%", ln=True)
        pdf.cell(0, 10, f"Preventivas canceladas: {cancelados}", ln=True)

        pdf.ln(5)
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 10, "Ranking de Cancelamentos:", ln=True)
        pdf.set_font("Arial", size=11)
        for motivo, qtd in frame_ranking.ranking.most_common():
            pdf.cell(0, 10, f"- {motivo}: {qtd} vez(es)", ln=True)

        pdf.ln(5)
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 10, "Detalhes dos Arquivos Cancelados:", ln=True)
        pdf.set_font("Arial", size=10)
        for nome, motivo, obs in frame_ranking.lista_detalhada:
            pdf.multi_cell(0, 8, f"Arquivo: {nome}\nMotivo: {motivo}\nObservações: {obs}\n", border=0)

        # Inserir gráfico
        pdf.image(fig_path, x=10, w=180)

        os.remove(fig_path)
        pdf.output(caminho)
        messagebox.showinfo("Exportação", f"PDF exportado com sucesso para:\n{caminho}")

    janela = tk.Toplevel()
    janela.title("Estatísticas Comerciais")
    janela.geometry("820x950")
    janela.configure(bg="#f4f4f4")
    janela.grab_set()

    pasta_confirmados = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\orçamentos\confirmados"
    pasta_concluidos = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\orçamentos\concluídos"
    pasta_pendentes = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\orçamentos\pendentes"
    pasta_cancelados = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\orçamentos\cancelados"

    frame_datas = tk.Frame(janela, bg="#f4f4f4")
    frame_datas.pack(pady=10)
    ttk.Label(frame_datas, text="Data Início (dd/mm/yyyy):").pack(side="left")
    entry_inicio = ttk.Entry(frame_datas)
    entry_inicio.pack(side="left", padx=5)
    entry_inicio.insert(0, "01/01/2025")
    ttk.Label(frame_datas, text="Data Fim:").pack(side="left")
    entry_fim = ttk.Entry(frame_datas)
    entry_fim.pack(side="left", padx=5)
    entry_fim.insert(0, datetime.now().strftime("%d/%m/%Y"))
    ttk.Button(frame_datas, text="Atualizar", command=carregar_dados).pack(side="left", padx=5)

    frame_info = tk.Frame(janela, bg="#f4f4f4", padx=20, pady=10)
    frame_info.pack(fill="x")

    fig = Figure(figsize=(8, 4), dpi=100)
    canvas = FigureCanvasTkAgg(fig, master=janela)
    canvas.get_tk_widget().pack(pady=10)

    frame_ranking = tk.Frame(janela, bg="#f4f4f4", padx=20, pady=10)
    frame_ranking.pack(fill="x")

    ttk.Button(janela, text="Exportar gráfico e dados para PDF", command=exportar_pdf).pack(pady=10)

    carregar_dados()

def gerar_tela_estatisticas_pecas():
    def contar_pecas_preventiva(diretorio):
        total = 0
        contagem = {}

        for arquivo in os.listdir(diretorio):
            if arquivo.endswith('.xlsx'):
                caminho = os.path.join(diretorio, arquivo)
                try:
                    df = pd.read_excel(caminho, engine='openpyxl')
                    df = df.dropna(how='all')
                    linhas = df.iloc[1:]

                    for _, row in linhas.iterrows():
                        nome = str(row[0]).strip()   # Coluna A
                        quantidade = row[1]          # Coluna B
                        codigo = str(row[2]).strip() # Coluna C

                        if pd.notna(codigo) and pd.notna(quantidade):
                            if codigo not in contagem:
                                contagem[codigo] = {'nome': nome, 'quantidade': 0}
                            contagem[codigo]['quantidade'] += int(quantidade)
                            total += int(quantidade)
                except Exception as e:
                    print(f'[ERRO] Falha ao processar {arquivo}: {e}')
                    continue

        ranking = sorted(contagem.items(), key=lambda x: x[1]['quantidade'], reverse=True)
        return total, ranking

    def contar_pecas_avulsas(html_path):
        total = 0
        contagem = defaultdict(lambda: {'nome': '', 'quantidade': 0})

        try:
            dfs = pd.read_html(html_path, skiprows=4)
            df = dfs[0]

            for _, row in df.iterrows():
                try:
                    grupo = str(row[7]).strip()  # Coluna H
                    if grupo != "22 - PEÇAS REVENDA":
                        continue

                    codigo = str(row[8]).strip()   # Coluna I
                    nome = str(row[9]).strip()     # Coluna J
                    quantidade = int(row[11])      # Coluna L

                    contagem[codigo]['nome'] = nome
                    contagem[codigo]['quantidade'] += quantidade
                    total += quantidade
                except Exception as e:
                    print(f'[ERRO] Linha com problema: {e}')
        except Exception as e:
            print(f'[ERRO] Leitura do HTML falhou: {e}')

        ranking = sorted(contagem.items(), key=lambda x: x[1]['quantidade'], reverse=True)
        return total, ranking

    def exibir_ranking(total_prev, ranking_prev, total_avul, ranking_avul):
        janela = tk.Toplevel()
        janela.title("Estatísticas - Ranking Completo")
        janela.geometry("1100x650")
        janela.configure(bg="#f4f4f4")
        janela.grab_set()

        ttk.Label(janela, text=f"Total de peças em preventivas: {total_prev}", font=("Arial", 12)).pack(pady=5)
        ttk.Label(janela, text=f"Total de peças avulsas: {total_avul}", font=("Arial", 12)).pack(pady=5)

        def criar_tabela_com_scroll(frame, titulo, dados):
            ttk.Label(frame, text=titulo, font=("Arial", 11, "bold")).pack()

            container = ttk.Frame(frame)
            container.pack(expand=True, fill="both", pady=10)

            canvas = tk.Canvas(container, height=500)
            scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
            scroll_frame = ttk.Frame(canvas)

            scroll_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(
                    scrollregion=canvas.bbox("all")
                )
            )

            canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)

            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")

            colunas = ("Código", "Nome da Peça", "Quantidade")
            tabela = ttk.Treeview(scroll_frame, columns=colunas, show="headings")
            for col in colunas:
                tabela.heading(col, text=col)
                tabela.column(col, width=280, anchor="center")

            for codigo, info in dados:
                tabela.insert("", "end", values=(codigo, info['nome'], info['quantidade']))

            tabela.pack(fill="both", expand=True)

        frame_rankings = ttk.Frame(janela)
        frame_rankings.pack(fill="both", expand=True, pady=10)

        frame1 = ttk.Frame(frame_rankings)
        frame1.pack(side="left", fill="both", expand=True, padx=10)

        frame2 = ttk.Frame(frame_rankings)
        frame2.pack(side="right", fill="both", expand=True, padx=10)

        criar_tabela_com_scroll(frame1, "Peças em Preventivas", ranking_prev)
        criar_tabela_com_scroll(frame2, "Peças Avulsas (Revenda)", ranking_avul)

    def exibir_grafico(titulo, dados):
        janela = tk.Toplevel()
        janela.title(f"Estatísticas - {titulo}")
        janela.geometry("1000x600")

        fig, ax = plt.subplots(figsize=(10, 6))
        codigos = [f"{c[0]}" for c in dados[:10]]
        quantidades = [c[1]['quantidade'] for c in dados[:10]]

        ax.barh(codigos[::-1], quantidades[::-1], color="#3498db")
        ax.set_xlabel("Quantidade")
        ax.set_ylabel("Código da Peça")
        ax.set_title(f"Top 10 {titulo}")

        canvas = FigureCanvasTkAgg(fig, master=janela)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def selecionar_visualizacao():
        root = tk.Tk()
        root.title("Escolher Visualização")
        root.geometry("300x180")

        ttk.Label(root, text="Selecione a forma de exibição:", font=("Arial", 11)).pack(pady=15)

        def mostrar_ranking():
            root.destroy()
            exibir_ranking(total_prev, ranking_prev, total_avul, ranking_avul)

        def mostrar_graficos():
            root.destroy()
            exibir_grafico("Peças Preventivas", ranking_prev)
            exibir_grafico("Peças Avulsas", ranking_avul)

        ttk.Button(root, text="Ranking Completo", command=mostrar_ranking).pack(pady=5)
        ttk.Button(root, text="Gráfico (Top 10)", command=mostrar_graficos).pack(pady=5)

        root.mainloop()

    # Caminhos
    caminho_preventivas = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\orçamentos\concluídos"
    caminho_html_avulsas = r"S:\ASSISTÊNCIA TÉCNICA\0-Luís Cappeletti\programa preventivas\dashboard\planilhas\historico_de_vendas\historico_vendas.xls"

    # Coleta de dados
    total_prev, ranking_prev = contar_pecas_preventiva(caminho_preventivas)
    total_avul, ranking_avul = contar_pecas_avulsas(caminho_html_avulsas)

    # Interface de escolha
    selecionar_visualizacao()

def abrir_configuracoes():
    import tkinter as tk
    from tkinter import messagebox, filedialog
    from datetime import datetime
    import shutil
    import os

    caminho_opcionais = "S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\personalizacao\\opcionais.txt"
    caminho_template = "S:\\ASSISTÊNCIA TÉCNICA\\0-Luís Cappeletti\\programa preventivas\\dashboard\\personalizacao\\template.pdf"
    pasta_template = os.path.dirname(caminho_template)

    # Tenta carregar os dias atuais do arquivo
    dias_1 = "1"
    dias_2 = "1"
    try:
        with open(caminho_opcionais, "r") as f:
            linhas = f.readlines()
            if len(linhas) >= 1:
                dias_1 = linhas[0].strip()
            if len(linhas) >= 2:
                dias_2 = linhas[1].strip()
    except:
        pass

    # Criar janela de configurações
    janela = tk.Toplevel()
    janela.title("⚙️ Configurações")
    janela.geometry("480x330")
    janela.configure(bg="white")
    janela.grab_set()

    fonte_padrao = ("Arial", 11)

    # Campo: Dias para aviso após preventiva
    tk.Label(janela, text="Dias para aviso após uma preventiva ser concluída:", bg="white", font=fonte_padrao).pack(pady=(15, 5))
    entrada1 = tk.Entry(janela, font=fonte_padrao, width=10, justify="center")
    entrada1.insert(0, dias_1)
    entrada1.pack()

    # Campo: Dias para cancelamento de orçamento
    tk.Label(janela, text="Dias para sugestão de cancelamento de orçamento:", bg="white", font=fonte_padrao).pack(pady=(15, 5))
    entrada2 = tk.Entry(janela, font=fonte_padrao, width=10, justify="center")
    entrada2.insert(0, dias_2)
    entrada2.pack()

    # Botão de trocar template
    def trocar_template_pdf():
        novo_arquivo = filedialog.askopenfilename(
            title="Selecione o novo template PDF",
            filetypes=[("Arquivos PDF", "*.pdf")]
        )
        if not novo_arquivo:
            return  # Cancelado

        try:
            if os.path.exists(caminho_template):
                data_suffix = datetime.now().strftime("%d%m%y")
                novo_nome = f"template{data_suffix}.pdf"
                caminho_backup = os.path.join(pasta_template, novo_nome)
                os.rename(caminho_template, caminho_backup)

            shutil.copy(novo_arquivo, caminho_template)
            messagebox.showinfo("Sucesso", "Template PDF trocado com sucesso!")

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao trocar o template: {e}")

    # Botão de salvar configurações
    def salvar_configuracoes():
        novo_dias1 = entrada1.get().strip()
        novo_dias2 = entrada2.get().strip()

        if not (novo_dias1.isdigit() and novo_dias2.isdigit()):
            messagebox.showwarning("Entrada inválida", "Digite apenas números inteiros.")
            return

        try:
            with open(caminho_opcionais, "w") as f:
                f.write(novo_dias1 + "\n")
                f.write(novo_dias2 + "\n")
            messagebox.showinfo("Sucesso", "Configurações salvas com sucesso.")
            janela.destroy()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar configurações: {e}")

    # Botão de trocar template
    tk.Button(janela, text="Trocar Template PDF", command=trocar_template_pdf, font=fonte_padrao, bg="#f4f4f4", fg="black").pack(pady=(25, 10))

    # Botão salvar
    tk.Button(janela, text="Salvar", command=salvar_configuracoes, font=fonte_padrao).pack(pady=5)


def main():
    global root
    root = tk.Tk()
    root.title("Dashboard Preventivas")
    root.geometry("1100x850")
    root.configure(bg="#f4f4f4")  # Cor de fundo para um layout mais agradável

    style = ttk.Style()
    style.theme_use("clam")



    # ========== FRAME PRINCIPAL ==========
    frame_principal = tk.Frame(root, padx=10, pady=10, bg="#f4f4f4")
    frame_principal.pack(fill="both", expand=True)

    # ========== SEÇÃO DE BOTÕES ==========
    frame_botoes = tk.Frame(frame_principal, bg="#f4f4f4")
    frame_botoes.pack(side="left", padx=10, pady=10, fill="y")

    botoes_info = [
        ("📌 Cadastros", [
            ("🧩 Cadastro de Peças", abrir_tela_cadastro_pecas),
            ("➕ Cadastro de Clientes", abrir_tela_cadastro_clientes),
            ("📅 Cadastro de Orçamentos", abrir_tela_orcamento_preventivas),
            ("📅 Cadastro de Preventivas", cadastro_preventivas),
        ]),
        ("🔍 Consultas", [

            ("🔍 Consulta Preventivas", tela_consulta_preventivas),
            ("🔍 Orçamentos pendentes", consultar_orcamentos_pendentes),
            ("🔍 Orçamentos cancelados", consultar_orcamentos_cancelados),
        ]),
        ("⚙️ Ações Gerais", [
            ("💡 Adicionar Lembrete", adicionar_lembrete),
            ("⚙️ Configurações", abrir_configuracoes),
        ])
    ]

    for titulo, botoes in botoes_info:
        frame_secao = ttk.LabelFrame(frame_botoes, text=titulo, padding=10)
        frame_secao.pack(fill="x", pady=5)

        for texto, comando in botoes:
            btn = ttk.Button(frame_secao, text=texto, command=comando, width=25)
            btn.pack(pady=3)

    # ========== SEÇÃO DE NOTIFICAÇÕES ==========
    frame_notificacoes = ttk.LabelFrame(frame_principal, text="⚠️ Vencimentos Próximos", padding=10)
    frame_notificacoes.pack(side="left", fill="y", padx=10)

    canvas = tk.Canvas(frame_notificacoes, height=200, width=250, bg="white")
    scroll_y = ttk.Scrollbar(frame_notificacoes, orient="vertical", command=canvas.yview)

    notificacoes_frame = tk.Frame(canvas, bg="white")
    canvas.create_window((0, 0), window=notificacoes_frame, anchor="nw")
    canvas.configure(yscrollcommand=scroll_y.set)

    canvas.pack(side="left", fill="both", expand=True)
    scroll_y.pack(side="right", fill="y")

    # ========== SEÇÃO DA TABELA ==========
    frame_tabela = tk.Frame(frame_principal, bg="#f4f4f4")
    frame_tabela.pack(side="right", padx=10, pady=10, fill="both", expand=True)

    tree = ttk.Treeview(frame_tabela, columns=("Nome", "Início", "Fim", "Caminho"), show='headings', height=12)
    tree.heading("Nome", text="Cliente")
    tree.heading("Início", text="Data Início")
    tree.heading("Fim", text="Data Fim")
    tree.column("Caminho", width=0, stretch=tk.NO)
    tree.pack(fill="both", expand=True)

    # ========== SEÇÃO INFERIOR (CALENDÁRIO + GRÁFICOS) ==========
    frame_inferior = tk.Frame(root, pady=10, bg="#f4f4f4")
    frame_inferior.pack(fill="both", expand=True)

    frame_calendario = ttk.LabelFrame(frame_inferior, text="📅 Calendário", padding=10)
    frame_calendario.pack(side="left", padx=10, fill="y")

    calendar = Calendar(frame_calendario, selectmode='day')
    calendar.pack()

    frame_grafico1 = ttk.LabelFrame(frame_inferior, text="📊 Ações Rápidas", padding=10)
    frame_grafico1.pack(side="left", padx=10, fill="both", expand=True)

    botoes_rapidos = [
        ("📈 Estatísticas de cadastro", estatistica_geral),
        ("📈 Efetividade Comercial", estatisticas_comerciais),
        ("🔧 Peças e Vendas", gerar_tela_estatisticas_pecas),
        ("📅 Controle de Tempo", cadastro_preventivas),
        ("🚨 Alertas e Manutenção", adicionar_lembrete),
    ]

    for texto, comando in botoes_rapidos:
        btn = ttk.Button(frame_grafico1, text=texto, command=comando, width=30)
        btn.pack(pady=5)

    frame_grafico2 = ttk.LabelFrame(frame_inferior, text="📈 Gráfico 2", padding=10)
    frame_grafico2.pack(side="left", padx=10, fill="both", expand=True)

    atualizar_tickets(tree, calendar)
    tree.bind("<Double-Button-1>", lambda e: on_ticket_select(e, tree))
    verificar_notificacoes(notificacoes_frame, root)
    notificacoes_frame.update_idletasks()
    canvas.config(scrollregion=canvas.bbox("all"))

    mover_orcamentos_antigos()  # Chamada da função automática para mover orçamentos antigos pendentes para cancelados

    root.mainloop()


if __name__ == "__main__":
    garantir_estrutura_minima()
    converter_csv_para_xlsx()
    main()


